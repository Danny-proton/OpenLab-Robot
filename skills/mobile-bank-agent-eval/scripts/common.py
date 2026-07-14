#!/usr/bin/env python3
"""common.py — mobile-bank-agent-eval 共享工具。

提供：
- Excel 读写（openpyxl，无则 fallback CSV）
- LLM 调用（有 API key 调真实 LLM，无则 fallback mock）
- mock LLM 响应（无需 API key 即可跑通全流程）
- 配置加载
- trace 事件生成（UATR 兼容）
"""

from __future__ import annotations

import json
import os
import re
import sys
import time
import hashlib
from datetime import datetime
from pathlib import Path
from typing import Any


# ---------------------------------------------------------------------------
# Excel 读写
# ---------------------------------------------------------------------------

def write_excel(data: dict, path: str, sheets: list[tuple[str, list[list]]] | None = None) -> None:
    """写 Excel。data 格式: {sheet_name: [[row], [row], ...]}
    sheets 优先于 data。
    """
    try:
        from openpyxl import Workbook
    except ImportError:
        _write_csv_fallback(data, path)
        return

    wb = Workbook()
    first = True
    if sheets:
        for sheet_name, rows in sheets:
            if first:
                ws = wb.active
                ws.title = sheet_name
                first = False
            else:
                ws = wb.create_sheet(sheet_name)
            for row in rows:
                ws.append(row)
    elif isinstance(data, dict):
        for sheet_name, rows in data.items():
            if first:
                ws = wb.active
                ws.title = sheet_name
                first = False
            else:
                ws = wb.create_sheet(sheet_name)
            for row in rows:
                ws.append(row)

    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def read_excel(path: str) -> dict[str, list[dict]]:
    """读 Excel，返回 {sheet_name: [{col: val}, ...]}。"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return _read_csv_fallback(path)

    wb = load_workbook(path, read_only=True)
    result: dict[str, list[dict]] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            result[sheet_name] = []
            continue
        headers = [str(h or "") for h in rows[0]]
        sheet_data = []
        for row in rows[1:]:
            if not row or not any(row):
                continue
            row_dict = {}
            for i, val in enumerate(row):
                if i < len(headers):
                    row_dict[headers[i]] = str(val) if val is not None else ""
            sheet_data.append(row_dict)
        result[sheet_name] = sheet_data
    return result


def _write_csv_fallback(data: dict, path: str) -> None:
    import csv
    base = path.replace(".xlsx", "")
    for sheet_name, rows in data.items():
        csv_path = f"{base}_{sheet_name}.csv"
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            for row in rows:
                w.writerow(row)
    print(f"[WARN] openpyxl 未安装，CSV fallback: {base}_*.csv", file=sys.stderr)


def _read_csv_fallback(path: str) -> dict[str, list[dict]]:
    import csv
    base = path.replace(".xlsx", "")
    result = {}
    for p in Path(base).parent.glob(f"{Path(base).name}_*.csv"):
        sheet_name = p.stem.split("_")[-1]
        with open(p, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            result[sheet_name] = [dict(row) for row in reader]
    return result


# ---------------------------------------------------------------------------
# LLM 调用（含 mock fallback）
# ---------------------------------------------------------------------------

def call_llm(system_prompt: str, user_prompt: str, timeout: int = 300) -> str:
    """调 LLM。有 API key 调真实 LLM，无则用 mock。

    环境变量:
    - LLM_API_KEY: API key（有则调真实 LLM）
    - LLM_MODEL: 模型名（默认 gpt-4o）
    - LLM_BASE_URL: API 地址（默认 OpenAI）
    - LLM_TIMEOUT: 超时秒数
    """
    api_key = os.getenv("LLM_API_KEY", "")
    model = os.getenv("LLM_MODEL", "gpt-4o")
    base_url = os.getenv("LLM_BASE_URL", "https://api.openai.com/v1")

    if not api_key:
        # mock fallback
        print("[INFO] LLM_API_KEY 未设置，使用 mock LLM", file=sys.stderr)
        return mock_llm_response(system_prompt, user_prompt)

    try:
        import requests
        url = base_url.rstrip("/") + "/chat/completions"
        headers = {
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        }
        body = {
            "model": model,
            "messages": [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt},
            ],
            "temperature": 0.3,
        }
        resp = requests.post(url, headers=headers, json=body, timeout=timeout)
        resp.raise_for_status()
        return resp.json()["choices"][0]["message"]["content"]
    except Exception as e:
        print(f"[WARN] LLM 调用失败，fallback 到 mock: {e}", file=sys.stderr)
        return mock_llm_response(system_prompt, user_prompt)


def mock_llm_response(system_prompt: str, user_prompt: str) -> str:
    """mock LLM 响应。根据 prompt 内容生成合理的 JSON。

    判断是需求分析还是用例生成，返回不同 mock 数据。
    """
    if "需求" in system_prompt and "维度" in system_prompt:
        # 需求分析 mock
        mock_data = {
            "dimensions": [
                {"id": "DIM-001", "name": "账户查询场景覆盖", "type": "业务场景覆盖"},
                {"id": "DIM-002", "name": "转账支付流程覆盖", "type": "业务流程覆盖"},
                {"id": "DIM-003", "name": "用户角色与意图覆盖", "type": "用户角色与意图覆盖"},
                {"id": "DIM-004", "name": "业务规则与约束覆盖", "type": "业务规则与约束覆盖"},
                {"id": "DIM-005", "name": "输入形态与上下文覆盖", "type": "输入形态与上下文覆盖"},
                {"id": "DIM-006", "name": "安全与边界覆盖", "type": "安全与边界覆盖"},
                {"id": "DIM-007", "name": "多轮对话状态覆盖", "type": "业务流程覆盖"},
                {"id": "DIM-008", "name": "异常恢复流程覆盖", "type": "业务流程覆盖"},
                {"id": "DIM-009", "name": "性能与延迟边界覆盖", "type": "安全与边界覆盖"},
                {"id": "DIM-010", "name": "合规与监管覆盖", "type": "业务规则与约束覆盖"},
            ],
            "scenarios": [
                {"id": "SC-001", "dimension": "DIM-001", "name": "查询账户余额", "description": "用户登录后查询主账户余额"},
                {"id": "SC-002", "dimension": "DIM-001", "name": "查询交易明细", "description": "按时间范围查询交易记录"},
                {"id": "SC-003", "dimension": "DIM-002", "name": "正常转账", "description": "行内转账，金额充足"},
                {"id": "SC-004", "dimension": "DIM-002", "name": "跨行转账", "description": "跨行转账，含手续费"},
                {"id": "SC-005", "dimension": "DIM-003", "name": "VIP用户查询", "description": "VIP用户专属产品查询"},
                {"id": "SC-006", "dimension": "DIM-004", "name": "单笔限额校验", "description": "超过单笔限额时拦截"},
                {"id": "SC-007", "dimension": "DIM-005", "name": "错别字输入", "description": "用户输入含错别字"},
                {"id": "SC-008", "dimension": "DIM-006", "name": "提示注入攻击", "description": "恶意 prompt 注入"},
                {"id": "SC-009", "dimension": "DIM-007", "name": "多轮转账确认", "description": "转账前多轮确认流程"},
                {"id": "SC-010", "dimension": "DIM-008", "name": "网络超时恢复", "description": "请求超时后重试"},
                {"id": "SC-011", "dimension": "DIM-009", "name": "大金额延迟", "description": "大金额转账延迟检测"},
                {"id": "SC-012", "dimension": "DIM-010", "name": "适当性管理", "description": "风险评估后推荐产品"},
            ],
            "skill_suggestions": [
                {"dimension_id": "DIM-001", "dimension_name": "账户查询场景覆盖", "skill": "account-service", "reason": "账户相关场景"},
                {"dimension_id": "DIM-002", "dimension_name": "转账支付流程覆盖", "skill": "transfer-service", "reason": "转账流程场景"},
            ]
        }
        return json.dumps(mock_data, ensure_ascii=False, indent=2)

    elif "测试用例" in system_prompt or "test_case" in system_prompt.lower():
        # 用例生成 mock — 解析 user_prompt 里的场景
        try:
            # 提取场景列表（JSON 数组）
            m = re.search(r'\[.*\]', user_prompt, re.DOTALL)
            if m:
                scenarios = json.loads(m.group(0))
            else:
                scenarios = [{"id": "SC-001", "name": "默认场景", "dimension_id": "DIM-001"}]
        except (json.JSONDecodeError, AttributeError):
            scenarios = [{"id": "SC-001", "name": "默认场景", "dimension_id": "DIM-001"}]

        test_cases = []
        tc_counter = 0
        for sc in scenarios:
            sc_id = sc.get("id", f"SC-{tc_counter+1:03d}")
            sc_name = sc.get("name", "场景")
            dim_id = sc.get("dimension_id", sc.get("dimension", sc.get("所属维度", "DIM-001")))
            # 每个场景生成 2 条用例
            tc_counter += 1
            test_cases.append({
                "scenario_id": sc_id,
                "dimension_id": dim_id,
                "tc_id": f"TC-{tc_counter:04d}",
                "title": f"{sc_name}-正常流程",
                "priority": "高",
                "precondition": "用户已登录，账户状态正常",
                "steps": ["1. 发起请求", "2. 等待响应", "3. 验证结果"],
                "user_input": f"请帮我{sc_name}",
                "expected_result": "余额,查询",  # 简短关键词，contains 断言用
                "assertion_type": "contains",
            })
            tc_counter += 1
            test_cases.append({
                "scenario_id": sc_id,
                "dimension_id": dim_id,
                "tc_id": f"TC-{tc_counter:04d}",
                "title": f"{sc_name}-异常流程",
                "priority": "中",
                "precondition": "用户已登录，但数据异常",
                "steps": ["1. 发起异常请求", "2. 等待响应", "3. 验证错误处理"],
                "user_input": f"请帮我{sc_name}（参数异常）",
                "expected_result": "异常,重试",  # 简短关键词
                "assertion_type": "contains",
            })

        return json.dumps({"test_cases": test_cases}, ensure_ascii=False, indent=2)

    # 默认 mock
    return json.dumps({"result": "mock response"}, ensure_ascii=False)


def extract_json(raw: str) -> dict:
    """从 LLM 响应中提取 JSON。"""
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        pass
    # 尝试 markdown code block
    m = re.search(r"```(?:json)?\s*([\s\S]*?)```", raw)
    if m:
        try:
            return json.loads(m.group(1))
        except json.JSONDecodeError:
            pass
    # 尝试找第一个 { 到最后 }
    m = re.search(r"\{[\s\S]*\}", raw)
    if m:
        try:
            return json.loads(m.group(0))
        except json.JSONDecodeError:
            pass
    raise ValueError(f"无法从 LLM 响应中提取 JSON: {raw[:200]}")


# ---------------------------------------------------------------------------
# UATR trace 事件生成
# ---------------------------------------------------------------------------

def make_trace_event(
    run_id: str,
    case_id: str,
    case_run_id: str,
    event_type: str,
    step: int,
    tool: str = "",
    arguments: dict | None = None,
    result: str = "",
    status: str = "success",
    latency_ms: int = 0,
) -> dict:
    """生成 UATR 兼容的 trace 事件。"""
    return {
        "schema_version": "uatr-0.5",
        "run_id": run_id,
        "case_id": case_id,
        "case_run_id": case_run_id,
        "trace_id": f"trace-{case_run_id}",
        "span_id": f"span_{step:04d}",
        "timestamp": datetime.now().astimezone().isoformat(timespec="seconds"),
        "framework": "mobile-bank-eval",
        "source": "execute_testcases",
        "event_type": event_type,
        "actor": {"type": "agent", "name": "mobile-bank-agent", "role": "executor"},
        "component": {"type": "tool" if tool else "agent", "name": tool or "mobile-bank-agent"},
        "status": status,
        "metrics": {"latency_ms": latency_ms},
        "attributes": {"tool.arguments": arguments} if arguments else {},
        "output": {"summary": result[:200]} if result else {},
    }


# ---------------------------------------------------------------------------
# 配置
# ---------------------------------------------------------------------------

def load_config(config_path: str) -> dict:
    """加载 YAML 配置。"""
    try:
        import yaml
    except ImportError:
        print("[ERROR] PyYAML 未安装", file=sys.stderr)
        sys.exit(1)
    with open(config_path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f) or {}


def now_str() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def make_run_id(variant: str = "eval", label: str = "") -> str:
    ts = datetime.now().strftime("%Y%m%d-%H%M%S")
    parts = [ts, variant]
    if label:
        safe = re.sub(r"[^a-z0-9_-]", "", label.lower())[:32]
        if safe:
            parts.append(safe)
    return "-".join(parts)


__all__ = [
    "write_excel", "read_excel",
    "call_llm", "mock_llm_response", "extract_json",
    "make_trace_event",
    "load_config", "now_str", "make_run_id",
]
