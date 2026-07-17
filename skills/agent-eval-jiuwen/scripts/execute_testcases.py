"""
mobileAgentTest - 测试用例执行
读取阶段 2 产出的测试用例 Excel，根据用户提供的环境信息执行 HTTP 请求，输出执行结果。

Usage:
    python execute_testcases.py --input path/to/test_cases.xlsx --output path/to/results.xlsx
                                --base-url http://localhost:8080 [--method POST]
                                [--timeout 30] [--headers '{"Content-Type":"application/json"}']
                                [--body '{"messages":[{"role":"user","content":"{{请求输入}}"}]}']
                                [--cases TC-0001,TC-0002] [--stream]
"""
import sys
import json
import os
import re
import argparse
import time
import urllib.parse
import io


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", default=None, help="测试用例 Excel 文件路径")
    parser.add_argument("--output", default=None, help="输出文件路径")
    parser.add_argument("--base-url", default="", help="目标环境基础 URL")
    parser.add_argument("--method", default="POST", help="HTTP 请求方法（默认 POST）")
    parser.add_argument("--timeout", type=int, default=120, help="请求超时秒数")
    parser.add_argument("--headers", default='{"Content-Type": "application/json"}', help="请求头 JSON（示例：'{\"Content-Type\": \"application/json\", \"Authorization\": \"Bearer xxx\"}'）")
    parser.add_argument("--body", default="", help="请求体 JSON 模板，{{列名}} 会被替换为测试用例对应列的内容")
    parser.add_argument("--cases", default="", help="指定运行的用例 ID，逗号分隔（默认全部）")
    parser.add_argument("--stream", action="store_true", default=False, help="SSE 流式响应模式，逐行读取 data: 事件并累加内容")
    args, _ = parser.parse_known_args()

    script_dir = os.path.dirname(os.path.abspath(__file__))
    data_dir = os.path.join(script_dir, "..", "data")

    input_path = args.input or os.path.join(data_dir, "test_cases.xlsx")
    input_path = _resolve_path(input_path)
    data_output = os.path.join(data_dir, "execution_results.xlsx")
    output_path = args.output or data_output
    output_path = os.path.abspath(output_path)

    if not os.path.exists(input_path):
        print(f"[ERROR] 测试用例文件不存在: {input_path}", file=sys.stderr)
        sys.exit(1)

    if not args.base_url:
        print("[ERROR] 请提供 --base-url（目标环境 URL）", file=sys.stderr)
        sys.exit(1)

    import requests

    selected = [c.strip() for c in args.cases.split(",") if c.strip()] if args.cases else []
    extra_headers = {}
    if args.headers:
        try:
            extra_headers = json.loads(args.headers)
        except json.JSONDecodeError:
            headers_env = os.getenv("HEADERS_JSON", "")
            if headers_env:
                try:
                    extra_headers = json.loads(headers_env)
                    print(f"[OK] headers from HEADERS_JSON env var", file=sys.stderr)
                except json.JSONDecodeError:
                    print(f"[WARN] HEADERS_JSON 环境变量格式错误", file=sys.stderr)
            else:
                print(f"[WARN] headers 格式错误，使用默认", file=sys.stderr)

    test_cases, headers_list = _read_test_cases(input_path)
    if not test_cases:
        print("[ERROR] 未找到测试用例", file=sys.stderr)
        sys.exit(1)

    if selected:
        id_key = headers_list[0] if headers_list else "用例 ID"
        test_cases = [tc for tc in test_cases if tc.get(id_key, "") in selected]
        print(f"case filter: selected {len(test_cases)}", file=sys.stderr)

    print(f"test_cases: {len(test_cases)}, base_url: {args.base_url}", file=sys.stderr)

    headers = dict(extra_headers)

    body_template = args.body
    if body_template:
        try:
            json.loads(_replace_placeholders(body_template, {h: "" for h in headers_list}))
        except json.JSONDecodeError:
            body_template = ""
    if not body_template:
        body_template = os.getenv("BODY_JSON", "")
    if body_template:
        try:
            json.loads(_replace_placeholders(body_template, {h: "" for h in headers_list}))
        except json.JSONDecodeError as e:
            print(f"[WARN] body JSON 占位符替换后仍有格式问题: {e}", file=sys.stderr)
            print(f"[WARN] 可用列: {headers_list}", file=sys.stderr)
            print(f"[WARN] 将继续执行，若运行时失败将标记为阻塞", file=sys.stderr)

    results = []
    passed = 0
    failed = 0
    blocked = 0

    for i, tc in enumerate(test_cases):
        path, method = _parse_request_info(tc.get("steps", ""), tc.get("title", ""), args.method)
        clean_path = path.lstrip("/")
        url = args.base_url.rstrip("/") + ("/" + clean_path if clean_path else "")
        if body_template:
            method = "POST"
        start = time.time()
        req_body = None
        body_err = None
        if body_template:
            body_str = _replace_placeholders(body_template, tc)
            try:
                req_body = json.loads(body_str)
            except json.JSONDecodeError:
                body_err = body_str
                req_body = None
        stream_meta = {}
        try:
            if body_err:
                raise ValueError(f"请求体 JSON 格式错误: {body_err[:200]}")
            if args.stream:
                resp = requests.request(method, url, headers=headers, json=req_body, timeout=args.timeout, stream=True)
                elapsed = int((time.time() - start) * 1000)
                status = resp.status_code
                if status != 200:
                    resp_body = resp.text
                else:
                    resp_body, stream_meta = _read_sse_stream(resp, start)
                ok = status < 500
                result = "通过" if ok else "失败"
            else:
                resp = requests.request(method, url, headers=headers, json=req_body, timeout=args.timeout)
                elapsed = int((time.time() - start) * 1000)
                status = resp.status_code
                content_type = resp.headers.get("Content-Type", "")
                if "text/event-stream" in content_type:
                    resp_body, stream_meta = _read_sse_stream(resp, start)
                else:
                    resp_body = resp.text
                ok = resp.ok
                result = "通过" if ok else "失败"
            if ok:
                passed += 1
            else:
                failed += 1
        except requests.exceptions.Timeout:
            elapsed = args.timeout * 1000
            status = 0
            resp_body = ""
            result = "阻塞"
            blocked += 1
        except Exception as e:
            elapsed = 0
            status = 0
            resp_body = str(e)
            result = "阻塞"
            blocked += 1

        tc_id = tc.get(headers_list[0], "") if headers_list else ""
        tc_scenario = tc.get(headers_list[1], "") if len(headers_list) > 1 else ""
        tc_title = tc.get(headers_list[2], "") if len(headers_list) > 2 else ""
        row = {
            "tc_id": tc_id,
            "scenario": tc_scenario,
            "title": tc_title,
            "method": method,
            "url": url,
            "request_headers": json.dumps(headers, ensure_ascii=False),
            "request_body": json.dumps(req_body, ensure_ascii=False) if req_body else "",
            "status": status,
            "elapsed": f"{elapsed}ms",
            "response": resp_body[:500] if not args.stream else resp_body[:500],
            "result": result,
        }
        if args.stream and stream_meta:
            row["stream_mode"] = "SSE"
            row["stream_chunks"] = stream_meta.get("chunks", 0)
            row["stream_first_byte_ms"] = stream_meta.get("first_byte_ms", 0)
            row["stream_duration_ms"] = stream_meta.get("stream_duration_ms", 0)
            row["stream_has_complete"] = "是" if stream_meta.get("has_complete") else "否"
        else:
            row["stream_mode"] = ""
            row["stream_chunks"] = ""
            row["stream_first_byte_ms"] = ""
            row["stream_duration_ms"] = ""
            row["stream_has_complete"] = ""
        results.append(row)
        if (i + 1) % 5 == 0:
            print(f"progress: {i + 1}/{len(test_cases)}", file=sys.stderr)

    _write_results(results, output_path)
    if output_path != os.path.abspath(data_output):
        if os.path.exists(os.path.abspath(data_output)):
            os.remove(os.path.abspath(data_output))
        _write_results(results, os.path.abspath(data_output))
    print(f"results: {passed} passed, {failed} failed, {blocked} blocked", file=sys.stderr)
    print(f"[阶段 3/4] 测试执行完成")
    print(f"产出文件：data/execution_results.xlsx")
    print(f"")
    print(f"执行结果：共 {len(results)} 个用例，通过 {passed}，失败 {failed}，阻塞 {blocked}")


def _replace_placeholders(template: str, case: dict) -> str:
    """将模板中的 {{列名}} 替换为测试用例对应列的值"""
    def _replacer(m):
        key = m.group(1)
        return case.get(key, "")
    return re.sub(r"\{\{(.+?)\}\}", _replacer, template)


def _resolve_path(path):
    path = os.path.abspath(path)
    if os.path.exists(path):
        return path
    sess_out = os.getenv("SESSION_OUTPUT_DIR", "")
    if sess_out:
        alt = os.path.join(sess_out, "dataset", os.path.basename(path))
        if os.path.exists(alt):
            return alt
    return path


def _read_test_cases(input_path):
    from openpyxl import load_workbook
    wb = load_workbook(input_path, read_only=True)
    sheet_name = None
    for name in ("测试用例", "Sheet1"):
        if name in wb.sheetnames:
            sheet_name = name
            break
    if not sheet_name:
        return [], []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [str(h or "") for h in rows[0]]
    cases = []
    for row in rows[1:]:
        if not row or not row[0]:
            continue
        case = {}
        for i, val in enumerate(row):
            if i < len(headers):
                case[headers[i]] = str(val or "")
        cases.append(case)
    return cases, headers


def _parse_request_info(steps_text, title, default_method="POST"):
    """从测试步骤或标题中提取 HTTP 方法和路径（忽略中文内容）"""
    text = steps_text + " " + title
    m = re.search(r'(GET|POST|PUT|DELETE|PATCH|HEAD|OPTIONS)\s+(/\S+)', text, re.IGNORECASE)
    if m:
        return m.group(2), m.group(1).upper()
    m = re.search(r'(?:https?://[^/\s]+)?(/[\w\-./]+)', text)
    if m:
        return m.group(1), default_method
    return "/", default_method


def _read_sse_stream(response, start_time: float) -> tuple[str, dict]:
    """读取 SSE 流式响应，累加 CHUNK 的 content 字段。
    返回 (accumulated_content, meta) 其中 meta 包含:
      - chunks: 总 chunk 数
      - first_byte_ms: 首字节延迟 (ms)
      - stream_duration_ms: 流式持续时长 (ms)
      - has_complete: 是否收到 COMPLETE 事件
      - terminal: 是否标记 terminal
    """
    accumulated = []
    chunks = 0
    first_byte = None
    last_byte = None
    has_complete = False
    terminal = False

    for line in response.iter_lines(decode_unicode=True):
        if not line:
            continue
        now = time.time()
        if first_byte is None:
            first_byte = now
        last_byte = now

        line = line.strip()
        if not line.startswith("data:"):
            continue
        payload = line[5:].strip()
        if payload == "[DONE]":
            has_complete = True
            continue
        try:
            evt = json.loads(payload)
        except json.JSONDecodeError:
            continue
        evt_type = evt.get("type", "")
        if evt_type == "CHUNK":
            content = evt.get("content", "")
            if content:
                accumulated.append(content)
                chunks += 1
        if evt.get("complete") is True or evt.get("omplete") is True or evt_type == "COMPLETE":
            has_complete = True
        if evt.get("terminal") is True or evt.get("terminal") == "true":
            terminal = True

    meta = {
        "chunks": chunks,
        "first_byte_ms": int((first_byte - start_time) * 1000) if first_byte else 0,
        "stream_duration_ms": int((last_byte - first_byte) * 1000) if first_byte and last_byte else 0,
        "has_complete": has_complete,
        "terminal": terminal,
    }
    return "".join(accumulated), meta


def _write_results(results: list, path: str):
    if os.path.exists(path):
        os.remove(path)
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "执行结果"
    ws.append(["用例 ID", "场景引用", "标题", "方法", "URL", "请求头", "请求体", "状态码", "响应时间", "响应体(前500字符)", "结果", "流模式", "Chunk数", "首字节(ms)", "流时长(ms)", "COMPLETE"])
    for r in results:
        ws.append([
            r["tc_id"], r["scenario"], r["title"],
            r["method"], r["url"], r.get("request_headers", ""),
            r.get("request_body", ""), r["status"], r["elapsed"],
            r["response"], r["result"],
            r.get("stream_mode", ""), r.get("stream_chunks", ""),
            r.get("stream_first_byte_ms", ""), r.get("stream_duration_ms", ""),
            r.get("stream_has_complete", ""),
        ])
    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        import traceback
        traceback.print_exc()
        sys.exit(1)
