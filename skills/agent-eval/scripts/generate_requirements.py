"""
手机银行智能体评测 - 需求分析
读取需求描述，调用 LLM 生成测试维度与场景，输出到 Excel 文件。

正向分析:
    python generate_requirements.py --description "需求描述文本" --output output.xlsx
逆向分析:
    python generate_requirements.py --input case.json --output output.xlsx
"""
import sys
import json
import os
import re
import argparse
import csv
import requests


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--description", default="", help="需求描述文本（正向分析，多行用\\n分隔）")
    parser.add_argument("--input", default="", help="输入文件路径（正向/逆向通用）")
    parser.add_argument("--output", default=None, help="输出文件路径")
    parser.add_argument("--reverse", action="store_true", help="逆向分析模式（从测试用例反推）")
    parser.add_argument("--list", default=None, help="列出指定 Excel 中的维度信息")
    args = parser.parse_args()

    if args.list:
        _list_dimensions(args.list)
        return

    description = ""
    test_cases = ""
    is_reverse = args.reverse

    if args.input:
        input_file = _resolve_path(args.input)
        with open(input_file, "r", encoding="utf-8") as f:
            content = f.read()
        if is_reverse:
            test_cases = content
        else:
            description = content
    else:
        raw = args.description
        if "\\n" in raw:
            raw = raw.replace("\\n", "\n")
        description = raw

    if not description and not test_cases:
        print("[ERROR] 请提供 --description（正向分析）或 --input --reverse（逆向分析）", file=sys.stderr)
        sys.exit(1)

    api_key = os.getenv("LLM_API_KEY", "")
    model = os.getenv("LLM_MODEL", "gpt-4o")
    base_url = os.getenv("LLM_BASE_URL", "https://api.openai.com/v1")
    timeout = int(os.getenv("LLM_TIMEOUT", "300"))

    if not api_key:
        print("[ERROR] LLM_API_KEY not set", file=sys.stderr)
        sys.exit(1)

    if is_reverse:
        data = _call_reverse_analysis(api_key, model, base_url, test_cases, timeout)
    else:
        data = _call_forward_analysis(api_key, model, base_url, description, timeout)

    script_dir = os.path.dirname(os.path.abspath(__file__))
    data_output = os.path.join(script_dir, "..", "data", "requirements_analysis.xlsx")
    output_path = args.output
    if not output_path:
        output_path = data_output
    output_path = os.path.abspath(output_path)

    try:
        if os.path.exists(output_path):
            os.remove(output_path)
        _write_excel(data, output_path)
        if output_path != os.path.abspath(data_output):
            if os.path.exists(os.path.abspath(data_output)):
                os.remove(os.path.abspath(data_output))
            _write_excel(data, os.path.abspath(data_output))
        dims = data.get("dimensions", [])
        scs = data.get("scenarios", [])
        saved_size = os.path.getsize(output_path) if os.path.exists(output_path) else 0
        print(f"[阶段 1/4] 需求分析完成", file=sys.stderr)
        print(f"产出文件：{output_path} ({saved_size} bytes)", file=sys.stderr)
        print(f"dimensions: {len(dims)}, scenarios: {len(scs)}", file=sys.stderr)
        # stdout 输出完整分析结果，Agent 直接使用
        print(f"[阶段 1/4] 需求分析完成")
        print(f"产出文件：data/requirements_analysis.xlsx")
        print(f"")
        print(f"共 {len(dims)} 个测试维度，{len(scs)} 个测试场景：")
        for d in dims:
            print(f"  • {d.get('id','')} {d.get('name','')}（{d.get('type','')}）")
        sc_sample = scs[:5]
        if sc_sample:
            print(f"  (前 {len(sc_sample)} 个场景)")
            for s in sc_sample:
                print(f"    - {s.get('id','')} {s.get('name','')}")
    except ImportError:
        csv_dir = output_path.replace(".xlsx", "")
        _write_csv(data, csv_dir)
        print(f"[OK] requirements analysis saved -> {csv_dir}_*.csv (Excel not available, CSV fallback)")


def _resolve_path(path):
    """解析文件路径，若不存在则尝试从 SESSION_OUTPUT_DIR/dataset/ 下查找。"""
    path = os.path.abspath(path)
    if os.path.exists(path):
        return path
    sess_out = os.getenv("SESSION_OUTPUT_DIR", "")
    if sess_out:
        alt = os.path.join(sess_out, "dataset", os.path.basename(path))
        if os.path.exists(alt):
            return alt
    return path


def _list_dimensions(path):
    path = _resolve_path(path)
    if not os.path.exists(path):
        print(f"[阶段 1/4] 需求分析完成", file=sys.stderr)
        print(f"产出文件：data/requirements_analysis.xlsx", file=sys.stderr)
        return
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True)
    if "测试维度" not in wb.sheetnames:
        return
    ws = wb["测试维度"]
    print("测试维度列表：")
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0]:
            print(f"  {row[0]} - {row[1]} ({row[2]})")


def _build_headers(api_key, base_url):
    return {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
    }


def _call_llm(api_key, model, base_url, system_prompt, user_prompt, timeout):
    url = base_url.rstrip("/") + "/chat/completions"
    headers = _build_headers(api_key, base_url)
    body = {
        "model": model,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ],
        "temperature": 0.3,
    }
    extra_opts = os.getenv("LLM_EXTRA_OPTS", "")
    if extra_opts:
        try:
            body.update(json.loads(extra_opts))
        except json.JSONDecodeError:
            pass

    try:
        resp = requests.post(url, headers=headers, json=body, timeout=timeout)
        resp.raise_for_status()
        content = resp.json()["choices"][0]["message"]["content"]
        return _extract_json(content)
    except Exception as e:
        print(f"[WARN] LLM call failed: {e}", file=sys.stderr)
        # 返回一个带错误标志的 fallback 结构
        return {
            "_error": str(e),
            "dimensions": [
                {"id": "ERR-001", "name": "LLM调用失败", "type": "异常"}
            ],
            "scenarios": [
                {"id": "SC-ERR", "dimension": "ERR-001",
                 "name": "默认场景", "description": f"LLM 调用异常: {e}"}
            ],
            "skill_suggestions": []
        }


def _call_forward_analysis(api_key, model, base_url, description, timeout):
    system_prompt = (
        "# Role: 资深智能体（Agent）系统架构师与 QA 专家。你精通 agent_evaluation.pdf 中的业务测试覆盖方法论。\n\n"
        "## Profile: 你精通智能体（Agent）系统的测试与设计。你熟悉单 Agent 及多 Agent（Multi-Agent）协作架构。你审视需求时，不把 Agent 视为简单的文本生成器，而是视其为一个具备“目标-感知-规划-记忆-执行”闭环的复杂动态系统。\n\n"
        "## Task: 请仔细阅读我提供的【Agent 需求说明】，找出其中隐藏的架构漏洞、死循环风险以及工程落地痛点。"
        "## Requirements: 参考以下金融领域测试覆盖框架（来自 agent_evaluation.pdf 第2章）：\n"
        "1. 业务场景覆盖 — 按业务功能拆解用户典型场景（账户服务、转账支付、存款、贷款、信用卡、投资理财、投诉售后等）\n"
        "2. 业务流程覆盖 — 正常路径、替代路径、异常路径\n"
        "3. 用户角色与意图覆盖 — 身份差异、意图多样性（查询、操作、投诉、闲聊等）\n"
        "4. 业务规则与约束覆盖 — if-then 逻辑、数值约束、合规要求（如适当性管理、禁止刚性兑付）\n"
        "5. 输入形态与上下文覆盖 — 不完整信息、错别字、指代消解、多轮上下文\n"
        "6. 安全与边界覆盖 — 敏感信息泄露、越权操作、提示注入、诱导突破\n\n"
        "## 具体要求：\n"
        "- 每个测试维度对应上述6个覆盖类型之一（业务场景/业务流程/用户角色/业务规则/输入形态/安全边界）\n"
        "- 维度名称应体现具体业务场景（如「账户服务场景覆盖」而非仅「业务场景覆盖」）\n"
        "- 场景描述需结合需求中的具体功能点\n\n"
        "请严格按照 JSON 格式输出，不要包含 markdown 标记。"
    )
    user_prompt = f"""请基于以下需求描述，生成测试维度和测试场景。

需求描述：
{description}

请按以下 JSON 格式输出（只输出 JSON，不要 markdown 标记）：
{{
  "dimensions": [
    {{"id": "DIM-001", "name": "维度名称", "type": "覆盖类型"}}
  ],
  "scenarios": [
    {{"id": "SC-001", "dimension": "DIM-001", "name": "子场景", "description": "描述"}}
  ],
  "skill_suggestions": [
    {{"dimension_id": "DIM-001", "dimension_name": "维度名称", "skill": "所属Skill", "reason": "理由"}}
  ]
}}

要求：
1. 维度名称应结合具体业务场景命名，如「产品筛选场景覆盖」「产品解读意图覆盖」「筛选业务规则覆盖」等
2. 覆盖类型从以下6种中选择：业务场景覆盖、业务流程覆盖、用户角色与意图覆盖、业务规则与约束覆盖、输入形态与上下文覆盖、安全与边界覆盖
3. 每个维度至少包含 2-3 个具体场景
4. 场景描述要具体，体现前置条件或触发条件
"""
    return _call_llm(api_key, model, base_url, system_prompt, user_prompt, timeout)


def _call_reverse_analysis(api_key, model, base_url, test_cases, timeout):
    system_prompt = (
        "你是一个专业的需求分析工程师。你擅长从已有的测试用例集中逆向分析，"
        "反推出测试维度和场景，并判断各维度归属于哪个 Skill。"
        "请严格按照 JSON 格式输出，不要包含 markdown 标记。"
    )
    user_prompt = f"""请对以下测试用例集进行逆向分析。

测试用例：
{test_cases[:8000]}

请执行以下步骤：
1. 解析用例结构，按子智能体、意图类型、交互轮次、边界类型等维度对用例聚类
2. 从聚类结果反向归纳出测试维度
3. 判断每个维度应归属到哪个现有 Skill（orchestrator / requirements-analysis / test-case-generator / test-executor / test-reporter）或建议新建

请按以下 JSON 格式输出（只输出 JSON，不要 markdown 标记）：
{{
  "dimensions": [
    {{"id": "DIM-001", "name": "维度名称", "type": "覆盖类型"}}
  ],
  "scenarios": [
    {{"id": "SC-001", "dimension": "DIM-001", "name": "子场景", "description": "描述"}}
  ],
  "skill_suggestions": [
    {{"dimension_id": "DIM-001", "dimension_name": "维度名称", "skill": "所属Skill", "reason": "理由"}}
  ]
}}
"""
    return _call_llm(api_key, model, base_url, system_prompt, user_prompt, timeout)


def _extract_json(raw: str):
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        pass
    m = re.search(r"```(?:json)?\s*([\s\S]*?)```", raw)
    if m:
        try:
            return json.loads(m.group(1))
        except json.JSONDecodeError:
            pass
    m = re.search(r"\{[\s\S]*\}", raw)
    if m:
        try:
            return json.loads(m.group(0))
        except json.JSONDecodeError:
            pass
    print("[ERROR] Failed to parse JSON from LLM response", file=sys.stderr)
    print(f"[DEBUG] Raw response (first 500 chars): {raw[:500]}", file=sys.stderr)
    sys.exit(1)


def _write_excel(data: dict, path: str):
    from openpyxl import Workbook

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "测试维度"
    ws1.append(["维度 ID", "维度名称", "覆盖类型"])
    for dim in data.get("dimensions", []):
        ws1.append([dim.get("id", ""), dim.get("name", ""), dim.get("type", "")])

    ws2 = wb.create_sheet("测试场景")
    ws2.append(["场景 ID", "所属维度", "子场景", "描述"])
    for sc in data.get("scenarios", []):
        ws2.append([
            sc.get("id", ""),
            sc.get("dimension", ""),
            sc.get("name", ""),
            sc.get("description", ""),
        ])

    suggestions = data.get("skill_suggestions", [])
    if suggestions:
        ws3 = wb.create_sheet("Skill归属建议")
        ws3.append(["维度 ID", "维度名称", "归属 Skill", "说明"])
        for s in suggestions:
            ws3.append([
                s.get("dimension_id", ""),
                s.get("dimension_name", ""),
                s.get("skill", ""),
                s.get("reason", ""),
            ])

    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)


def _write_csv(data: dict, base_path: str):
    os.makedirs(os.path.dirname(base_path), exist_ok=True)

    dim_path = base_path + "_dimensions.csv"
    with open(dim_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["维度 ID", "维度名称", "覆盖类型"])
        for dim in data.get("dimensions", []):
            w.writerow([dim.get("id", ""), dim.get("name", ""), dim.get("type", "")])

    sc_path = base_path + "_scenarios.csv"
    with open(sc_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["场景 ID", "所属维度", "子场景", "描述"])
        for sc in data.get("scenarios", []):
            w.writerow([
                sc.get("id", ""),
                sc.get("dimension", ""),
                sc.get("name", ""),
                sc.get("description", ""),
            ])

    suggestions = data.get("skill_suggestions", [])
    if suggestions:
        sk_path = base_path + "_skill_suggestions.csv"
        with open(sk_path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["维度 ID", "维度名称", "归属 Skill", "说明"])
            for s in suggestions:
                w.writerow([
                    s.get("dimension_id", ""),
                    s.get("dimension_name", ""),
                    s.get("skill", ""),
                    s.get("reason", ""),
                ])


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        import traceback
        traceback.print_exc()
        sys.exit(1)
