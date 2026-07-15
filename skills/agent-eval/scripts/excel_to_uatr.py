"""
excel_to_uatr.py — mobile-bank 4 阶段流水线 → agent-eval eval loop 桥接器

把 mobile-bank 分支阶段 3 产出的 execution_results.xlsx（连同阶段 1/2 的 Excel）
翻译成 agent-eval 主分支能消费的 UATR trace + cases YAML + run 记录 + 初步评分。

产出（写入 .agent-eval/ 标准目录）：
  - traces/<run_id>.jsonl     UATR 格式 trace（每条用例一组 11 类事件中的若干）
  - cases/<run_id>.yaml       case 定义（含 expected / business_rules，供 diagnoser/multi_judge）
  - runs/<run_id>.jsonl       run 记录（case_id / status / latency_ms / final_answer / trace_path）
  - scores/<run_id>.json      初步机械评分（task_success / output_schema / latency / hard_fail）

桥接器是纯机械格式转换，不调任何 LLM。诊断的"软评分"由后续 multi_judge.py（9 个 judge
agent，Claude 自己扮演）完成。

Usage:
    python excel_to_uatr.py \
        --requirements data/requirements_analysis.xlsx \
        --testcases data/test_cases.xlsx \
        --results data/execution_results.xlsx \
        --config .agent-eval/config.yaml \
        --variant baseline \
        --label "mobile-bank-20260715"

    # 只桥接不写 cases YAML（调试用）
    python excel_to_uatr.py --results data/execution_results.xlsx --no-cases
"""
import sys
import os
import json
import argparse
import datetime
import re
from pathlib import Path


def main():
    parser = argparse.ArgumentParser(
        description="mobile-bank Excel → agent-eval UATR 桥接器（纯机械转换，无 LLM）"
    )
    parser.add_argument("--requirements", default=None, help="requirements_analysis.xlsx 路径")
    parser.add_argument("--testcases", default=None, help="test_cases.xlsx 路径")
    parser.add_argument("--results", required=True, help="execution_results.xlsx 路径")
    parser.add_argument("--config", default=".agent-eval/config.yaml", help="agent-eval config 路径")
    parser.add_argument("--variant", default="baseline", help="run variant 标签")
    parser.add_argument("--label", default="", help="run label")
    parser.add_argument("--no-cases", action="store_true", help="不生成 cases YAML（仅 trace + run + score）")
    args = parser.parse_args()

    # 读三个 Excel
    req_path = _resolve_path(args.requirements) if args.requirements else None
    tc_path = _resolve_path(args.testcases) if args.testcases else None
    res_path = _resolve_path(args.results)

    for p, name in [(req_path, "需求分析"), (tc_path, "测试用例"), (res_path, "执行结果")]:
        if p is None:
            continue
        if not os.path.exists(p):
            print(f"[ERROR] {name}文件不存在: {p}", file=sys.stderr)
            sys.exit(1)

    dims, scenarios, dim_map = _read_requirements(req_path) if req_path else ([], [], {})
    test_cases = _read_testcases(tc_path) if tc_path else []
    results = _read_results(res_path)
    if not results:
        print("[ERROR] 执行结果为空", file=sys.stderr)
        sys.exit(1)

    # 解析 config.yaml 拿目录配置
    cfg = _load_config(args.config)
    traces_dir = _cfg_dir(cfg, "traces", ".agent-eval/traces")
    runs_dir = _cfg_dir(cfg, "runs", ".agent-eval/runs")
    scores_dir = _cfg_dir(cfg, "scores", ".agent-eval/scores")
    cases_dir = _cfg_dir(cfg, "cases", ".agent-eval/cases")
    for d in (traces_dir, runs_dir, scores_dir, cases_dir):
        os.makedirs(d, exist_ok=True)

    # 生成 run_id
    ts = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
    label = args.label or f"mobile-bank-{ts}"
    run_id = f"{ts}-{args.variant}-{label}"
    # run_id 里不能有空格
    run_id = re.sub(r"\s+", "-", run_id)

    # 按 tc_id 索引用例和结果
    tc_by_id = {tc["tc_id"]: tc for tc in test_cases}
    # 维度 → 场景 → 用例 反查表
    sc_by_id = {s["id"]: s for s in scenarios}

    # 1. 生成 UATR trace
    trace_path = os.path.join(traces_dir, f"{run_id}.jsonl")
    run_records = []
    n_ok = n_fail = n_err = 0
    latencies = []
    with open(trace_path, "w", encoding="utf-8") as ft:
        for r in results:
            # execute_testcases.py 写的 header: "用例 ID"(带空格) / "响应体(前500字符)" / "响应时间" / "状态码" / "结果"
            tc_id = (r.get("用例 ID") or r.get("用例ID") or r.get("case_id") or "")
            tc_id = str(tc_id).strip() if tc_id else ""
            tc = tc_by_id.get(tc_id, {})
            scenario_id = tc.get("scenario_id", "")
            sc = sc_by_id.get(scenario_id, {})
            dim_id = tc.get("dimension_id") or sc.get("dimension_id", "")
            case_id = tc_id or f"case_{len(run_records)+1:04d}"
            case_run_id = f"{run_id}::{case_id}"
            final_answer = (
                r.get("响应体(前500字符)") or r.get("实际响应") or r.get("response") or ""
            )
            final_answer = str(final_answer)[:4000] if final_answer else ""
            status_code = _to_int(r.get("状态码") or r.get("status_code"))
            latency_ms = _to_int(
                r.get("响应时间") or r.get("响应时间ms") or r.get("latency_ms")
            )
            result_str = str(r.get("结果") or r.get("result") or "").strip()
            err_msg = str(r.get("错误信息") or r.get("error") or "").strip()
            # 5xx 或网络错也算 error
            if not err_msg and status_code and status_code >= 500:
                err_msg = f"HTTP {status_code}"
            user_input = tc.get("user_input", "")

            status = "success"
            if err_msg or (status_code and status_code >= 500):
                status = "error"
                n_err += 1
            elif result_str in ("失败", "fail", "FAIL", "Fail"):
                status = "fail"
                n_fail += 1
            else:
                n_ok += 1
            if latency_ms:
                latencies.append(latency_ms)

            now = datetime.datetime.now().astimezone().isoformat()
            step = 1
            # agent_start
            ft.write(json.dumps({
                "run_id": run_id, "case_id": case_id, "case_run_id": case_run_id,
                "ts": now, "event": "agent_start", "step": step, "agent": "mobile-bank-sut",
            }, ensure_ascii=False) + "\n"); step += 1
            # prompt_rendered
            ft.write(json.dumps({
                "run_id": run_id, "case_id": case_id, "case_run_id": case_run_id,
                "ts": now, "event": "prompt_rendered", "step": step,
                "prompt_hash": "sha256:" + _short_hash(user_input), "model": "unknown",
            }, ensure_ascii=False) + "\n"); step += 1
            # model_call
            ft.write(json.dumps({
                "run_id": run_id, "case_id": case_id, "case_run_id": case_run_id,
                "ts": now, "event": "model_call", "step": step, "model": "unknown",
                "input_tokens": None, "output_tokens": None,
            }, ensure_ascii=False) + "\n"); step += 1
            # 若有错误，写 error 事件
            if err_msg:
                ft.write(json.dumps({
                    "run_id": run_id, "case_id": case_id, "case_run_id": case_run_id,
                    "ts": now, "event": "error", "step": step,
                    "error": {"type": "http_error", "message": err_msg},
                }, ensure_ascii=False) + "\n"); step += 1
            # agent_final
            ft.write(json.dumps({
                "run_id": run_id, "case_id": case_id, "case_run_id": case_run_id,
                "ts": now, "event": "agent_final", "step": step,
                "final_answer": final_answer,
            }, ensure_ascii=False) + "\n"); step += 1
            # agent_end
            ft.write(json.dumps({
                "run_id": run_id, "case_id": case_id, "case_run_id": case_run_id,
                "ts": now, "event": "agent_end", "step": step,
                "status": status, "latency_ms": latency_ms,
            }, ensure_ascii=False) + "\n")

            run_records.append({
                "case_id": case_id,
                "case_run_id": case_run_id,
                "tc_id": tc_id,
                "scenario_id": scenario_id,
                "dimension_id": dim_id,
                "status": status,
                "status_code": status_code,
                "latency_ms": latency_ms,
                "final_answer": final_answer,
                "trace_path": trace_path,
                "expected": tc.get("expected", ""),
                "assertion_type": tc.get("assertion_type", "contains"),
                "user_input": user_input,
                "result_str": result_str,
                "error": err_msg,
            })

    # 2. 写 runs/<run_id>.jsonl
    run_path = os.path.join(runs_dir, f"{run_id}.jsonl")
    with open(run_path, "w", encoding="utf-8") as fr:
        for rec in run_records:
            fr.write(json.dumps(rec, ensure_ascii=False) + "\n")

    # 3. 生成 cases YAML（供 diagnoser/multi_judge 消费）
    cases_path = None
    if not args.no_cases and test_cases:
        cases_path = os.path.join(cases_dir, f"{run_id}.yaml")
        _write_cases_yaml(cases_path, run_id, test_cases, sc_by_id, dim_map, results)

    # 4. 初步机械评分
    score = _compute_score(run_records, latencies)
    score_path = os.path.join(scores_dir, f"{run_id}.json")
    with open(score_path, "w", encoding="utf-8") as fs:
        json.dump(score, fs, ensure_ascii=False, indent=2)

    # stdout 摘要
    summary = {
        "run_id": run_id,
        "traces_file": trace_path,
        "runs_file": run_path,
        "scores_file": score_path,
        "cases_file": cases_path,
        "total_cases": len(run_records),
        "success": n_ok,
        "fail": n_fail,
        "error": n_err,
        "avg_latency_ms": round(sum(latencies) / len(latencies), 1) if latencies else None,
        "weighted_score": score.get("weighted_score"),
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    print(f"\n[桥接完成] run_id={run_id}", file=sys.stderr)
    print(f"  trace: {trace_path}", file=sys.stderr)
    print(f"  runs:  {run_path}", file=sys.stderr)
    print(f"  score: {score_path}", file=sys.stderr)
    if cases_path:
        print(f"  cases: {cases_path}", file=sys.stderr)
    print(f"  下一步: python scripts/diagnoser.py --config {args.config} --run {run_id}", file=sys.stderr)


# ---------- Excel 读取 ----------

def _resolve_path(path):
    path = os.path.abspath(path)
    if os.path.exists(path):
        return path
    sess_out = os.getenv("SESSION_OUTPUT_DIR", "")
    if sess_out:
        alt = os.path.join(sess_out, "dataset", os.path.basename(path))
        if os.path.exists(alt):
            return alt
    return path


def _read_requirements(path):
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True)
    dims = []
    dim_map = {}
    if "测试维度" in wb.sheetnames:
        ws = wb["测试维度"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                d = {"id": str(row[0]), "name": str(row[1] or ""), "type": str(row[2] or "")}
                dims.append(d)
                dim_map[d["id"]] = d
    scenarios = []
    if "测试场景" in wb.sheetnames:
        ws = wb["测试场景"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                scenarios.append({
                    "id": str(row[0]),
                    "dimension": str(row[1] or ""),
                    "name": str(row[2] or ""),
                    "description": str(row[3] or ""),
                })
    return dims, scenarios, dim_map


def _read_testcases(path):
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True)
    ws = wb["测试用例"] if "测试用例" in wb.sheetnames else wb.active
    cases = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        cases.append({
            "tc_id": str(row[0] or ""),
            "scenario_id": str(row[1] or ""),
            "dimension_id": str(row[2] or "") if len(row) > 2 else "",
            "title": str(row[3] or "") if len(row) > 3 else "",
            "priority": str(row[4] or "") if len(row) > 4 else "",
            "preconditions": str(row[5] or "") if len(row) > 5 else "",
            "steps": str(row[6] or "") if len(row) > 6 else "",
            "user_input": str(row[7] or "") if len(row) > 7 else "",
            "expected": str(row[8] or "") if len(row) > 8 else "",
            "assertion_type": str(row[9] or "contains") if len(row) > 9 else "contains",
        })
    return cases


def _read_results(path):
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h or "").strip() for h in rows[0]]
    results = []
    for row in rows[1:]:
        if not row or not row[0]:
            continue
        rec = {}
        for i, val in enumerate(row):
            if i < len(header):
                rec[header[i]] = val
        results.append(rec)
    return results


# ---------- 配置 ----------

def _load_config(path):
    if not os.path.exists(path):
        return {}
    try:
        import yaml
    except ImportError:
        return {}
    with open(path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f) or {}


def _cfg_dir(cfg, key, default):
    dirs = cfg.get("dirs", {}) if isinstance(cfg, dict) else {}
    rel = dirs.get(key, default)
    config_dir = os.path.dirname(os.path.abspath(os.path.normpath(cfg.get("_config_path", ".agent-eval/config.yaml")))) \
        if "_config_path" in cfg else ".agent-eval"
    # 相对路径基于 cwd
    if os.path.isabs(rel):
        return rel
    # config.yaml 里的 dirs 是相对 .agent-eval/ 的，但实际用法是相对 cwd
    # 主分支 eval_runner.py 把 .agent-eval/ 作为 base，dirs 相对 base
    base = ".agent-eval"
    return os.path.join(base, rel)


# ---------- cases YAML ----------

def _write_cases_yaml(path, run_id, test_cases, sc_by_id, dim_map, results):
    """生成 cases YAML，schema 对齐 examples/.agent-eval/cases/train.yaml"""
    # 按 tc_id 索引结果
    res_by_tc = {}
    for r in results:
        tc_id = (r.get("用例 ID") or r.get("用例ID") or r.get("case_id") or "")
        tc_id = str(tc_id).strip() if tc_id else ""
        if tc_id:
            res_by_tc[tc_id] = r

    lines = [f"# cases for run {run_id}", "# 由 excel_to_uatr.py 从 mobile-bank Excel 桥接生成", "", "cases:"]
    for tc in test_cases:
        tc_id = tc["tc_id"]
        sc = sc_by_id.get(tc.get("scenario_id"), {})
        res = res_by_tc.get(tc_id, {})
        actual = (res.get("响应体(前500字符)") or res.get("实际响应") or res.get("response") or "")
        actual = str(actual)[:1000] if actual else ""
        expected = tc.get("expected", "")
        atype = tc.get("assertion_type", "contains")
        lines.append(f"  - id: {tc_id}")
        lines.append(f"    name: {tc.get('title', '')}")
        lines.append(f"    agent: mobile-bank-sut")
        lines.append(f"    task: {sc.get('description', '')[:200]}")
        lines.append(f"    input:")
        lines.append(f"      user_message: {_yaml_str(tc.get('user_input', ''))}")
        if tc.get("preconditions"):
            lines.append(f"      preconditions: {_yaml_str(tc['preconditions'])}")
        lines.append(f"    expected:")
        if atype == "contains" and expected:
            lines.append(f"      final_decision:")
            lines.append(f"        contains:")
            for part in re.split(r"[;；\n]", expected):
                part = part.strip()
                if part:
                    lines.append(f"          - {_yaml_str(part)}")
        elif atype == "exact" and expected:
            lines.append(f"      final_decision:")
            lines.append(f"        equals: {_yaml_str(expected)}")
        elif atype == "regex" and expected:
            lines.append(f"      final_decision:")
            lines.append(f"        regex: {_yaml_str(expected)}")
        else:
            lines.append(f"      final_decision:")
            lines.append(f"        contains:")
            lines.append(f"          - {_yaml_str(expected)}")
        # business_rules 从维度 type 推导一个占位
        dim = dim_map.get(tc.get("dimension_id") or sc.get("dimension", ""), {})
        if dim:
            lines.append(f"    business_rules:")
            lines.append(f"      must_satisfy:")
            lines.append(f"        - id: {dim.get('id', 'dim_rule')}")
            lines.append(f"          description: {_yaml_str(dim.get('name', '') + ' / ' + dim.get('type', ''))}")
        lines.append(f"    meta:")
        lines.append(f"      scenario_id: {tc.get('scenario_id', '')}")
        lines.append(f"      dimension_id: {tc.get('dimension_id', '')}")
        lines.append(f"      priority: {tc.get('priority', '')}")
        lines.append(f"      actual_response: {_yaml_str(actual)}")
        lines.append("")
    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


def _yaml_str(s):
    """简单 YAML 字符串转义：含特殊字符则用双引号包"""
    s = str(s).strip()
    if not s:
        return '""'
    if any(c in s for c in [":", "#", "\n", '"', "'", "{", "}", "[", "]", ",", "&", "*", "!", "|", ">", "%", "@", "`"]):
        return '"' + s.replace('"', '\\"') + '"'
    return s


# ---------- 评分 ----------

def _compute_score(run_records, latencies):
    n = len(run_records)
    if n == 0:
        return {"weighted_score": 0.0, "n_total": 0}
    n_success = sum(1 for r in run_records if r["status"] == "success")
    n_fail = sum(1 for r in run_records if r["status"] == "fail")
    n_err = sum(1 for r in run_records if r["status"] == "error")
    task_success = n_success / n
    # output_schema: 响应非空且可解析为 JSON 的比例（粗略）
    n_schema_ok = 0
    for r in run_records:
        fa = r.get("final_answer", "")
        if fa:
            try:
                json.loads(fa)
                n_schema_ok += 1
            except Exception:
                # 非JSON但有内容也算半分
                if len(fa) > 10:
                    n_schema_ok += 0.5
    output_schema = n_schema_ok / n
    # latency：归一化到 [0,1]，500ms=1.0，5000ms=0.0
    if latencies:
        avg_lat = sum(latencies) / len(latencies)
        latency_score = max(0.0, min(1.0, (5000 - avg_lat) / 4500)) if avg_lat > 500 else 1.0
    else:
        latency_score = 0.0
    hard_fail = n_err  # 5xx / 网络错算硬失败
    # 加权（对齐 config.yaml weights，这里用默认）
    weighted = 0.35 * task_success + 0.20 * 0.5 + 0.20 * 0.5 + 0.15 * output_schema + 0.10 * latency_score
    weighted -= 0.5 * hard_fail / n  # 硬失败惩罚
    weighted = max(0.0, min(1.0, weighted))
    return {
        "run_id": run_records[0]["case_run_id"].split("::")[0] if run_records else "",
        "n_total": n,
        "n_success": n_success,
        "n_fail": n_fail,
        "n_error": n_err,
        "n_hard_fail": hard_fail,
        "task_success": round(task_success, 4),
        "output_schema": round(output_schema, 4),
        "latency_score": round(latency_score, 4),
        "avg_latency_ms": round(sum(latencies) / len(latencies), 1) if latencies else None,
        "weighted_score": round(weighted, 4),
        "scorer": "excel_to_uatr.mechanical",
        "note": "初步机械评分；软评分（F1-F8 / 9 Judge）由 diagnoser.py / multi_judge.py 后续补完",
    }


def _to_int(v):
    if v is None or v == "":
        return None
    try:
        return int(float(v))
    except Exception:
        return None


def _short_hash(s):
    import hashlib
    return hashlib.sha256(str(s).encode("utf-8")).hexdigest()[:12]


if __name__ == "__main__":
    try:
        main()
    except Exception:
        import traceback
        traceback.print_exc()
        sys.exit(1)
