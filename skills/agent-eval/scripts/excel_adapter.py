#!/usr/bin/env python3
"""excel_adapter.py — Excel 用例 → agent-eval YAML case 格式转换器。

把 mobileAgentTest 原版的 Excel 用例转成 agent-eval 的标准 YAML case 格式，
让 agent-eval 的 eval_runner 直接执行。

支持两种 Excel 格式：
1. mobileAgentTest 原版格式（用例ID/场景ID/维度ID/标题/优先级/前置条件/测试步骤/用户输入/预期结果/断言类型）
2. 通用格式（任意列名，自动映射）

用法:
  python excel_adapter.py --input test_cases.xlsx --output cases/train.yaml
  python excel_adapter.py --input test_cases.xlsx --output cases/train.yaml --mapping custom_mapping.json
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

try:
    import yaml
except ImportError:
    sys.stderr.write("[ERROR] PyYAML 未安装\n")
    sys.exit(1)


# 默认列名映射：Excel 列名 → agent-eval case 字段
DEFAULT_MAPPING = {
    "用例ID": "id",
    "用例id": "id",
    "tc_id": "id",
    "TC-ID": "id",
    "场景ID": "_scenario_id",
    "维度ID": "_dimension_id",
    "标题": "name",
    "优先级": "_priority",
    "前置条件": "_precondition",
    "测试步骤": "_steps",
    "用户输入": "input.user_message",
    "预期结果": "expected.final_decision.contains",
    "断言类型": "_assertion_type",
}


def convert_excel_to_cases(input_path: str, mapping: dict | None = None) -> list[dict]:
    """读 Excel，转成 agent-eval 格式的 case 列表。"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.stderr.write("[ERROR] openpyxl 未安装: pip install openpyxl\n")
        sys.exit(1)

    wb = load_workbook(input_path, read_only=True)
    mapping = mapping or DEFAULT_MAPPING

    # 找到包含数据的 sheet
    ws = None
    for name in wb.sheetnames:
        if "用例" in name or "test" in name.lower() or "case" in name.lower():
            ws = wb[name]
            break
    if not ws:
        ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h or "").strip() for h in rows[0]]
    cases = []

    for row in rows[1:]:
        if not row or not any(row):
            continue
        row_dict = {headers[i]: str(v).strip() if v is not None else ""
                    for i, v in enumerate(row)}

        case = _convert_row(row_dict, mapping)
        if case:
            cases.append(case)

    return cases


def _convert_row(row: dict, mapping: dict) -> dict | None:
    """把一行 Excel 数据转成 agent-eval case。"""
    case_id = ""
    user_input = ""
    expected_keywords = []

    # 找 id 和 user_input
    for excel_col, case_field in mapping.items():
        val = row.get(excel_col, "")
        if not val:
            continue

        if case_field == "id":
            case_id = val
        elif case_field == "input.user_message":
            user_input = val
        elif case_field == "expected.final_decision.contains":
            # 预期结果：逗号分隔的关键词
            expected_keywords = [k.strip() for k in val.split(",") if k.strip()]
        elif case_field == "_assertion_type":
            # 断言类型决定 expected 格式
            pass

    if not case_id:
        case_id = f"tc_{hash(str(row)) & 0xFFFF:04x}"

    # 构建 agent-eval case
    case = {
        "id": case_id,
        "name": row.get("标题", row.get("name", case_id)),
        "agent": "mobile-bank-agent",
        "task": row.get("标题", ""),
        "input": {
            "user_message": user_input or row.get("用户输入", ""),
        },
        "expected": {
            "final_decision": {
                "contains": expected_keywords if expected_keywords else [user_input[:10]],
            },
        },
        "expected_tools": {
            "required": [],
            "forbidden": [],
        },
        "business_rules": {
            "must_satisfy": [],
        },
        "expected_steps": 8,
        "scoring": {
            "hard_fail_if": ["forbidden_tool_called"],
        },
    }

    # 附加元数据（不影响 agent-eval 执行）
    if row.get("场景ID"):
        case["_scenario_id"] = row["场景ID"]
    if row.get("维度ID"):
        case["_dimension_id"] = row["维度ID"]
    if row.get("优先级"):
        case["_priority"] = row["优先级"]
    if row.get("前置条件"):
        case["_precondition"] = row["前置条件"]
    if row.get("测试步骤"):
        case["_steps"] = row["测试步骤"]
    if row.get("断言类型"):
        case["_assertion_type"] = row["断言类型"]

    return case


def main() -> int:
    ap = argparse.ArgumentParser(description="Excel → YAML case 转换器")
    ap.add_argument("--input", required=True, help="Excel 文件路径")
    ap.add_argument("--output", required=True, help="输出 YAML 路径")
    ap.add_argument("--mapping", default=None, help="自定义列名映射 JSON")
    args = ap.parse_args()

    mapping = None
    if args.mapping:
        with open(args.mapping, "r", encoding="utf-8") as f:
            mapping = json.load(f)

    cases = convert_excel_to_cases(args.input, mapping)

    output = {"cases": cases}
    Path(args.output).parent.mkdir(parents=True, exist_ok=True)
    with open(args.output, "w", encoding="utf-8") as f:
        yaml.safe_dump(output, f, allow_unicode=True, sort_keys=False)

    print(f"[OK] 转换完成: {args.input} → {args.output}")
    print(f"  用例数: {len(cases)}")
    for c in cases:
        print(f"  - {c['id']}: {c['name']}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
