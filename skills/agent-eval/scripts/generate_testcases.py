"""
mobileAgentTest - 测试用例生成
读取阶段 1 产出的需求分析 Excel，为每个场景生成详细测试用例，输出到 Excel 文件。

Usage:
    python generate_testcases.py --input path/to/requirements_analysis.xlsx --output path/to/test_cases.xlsx
"""
import sys
import json
import os
import re
import argparse
import requests


def _resolve_path(path):
    path = os.path.abspath(path)
    if os.path.exists(path):
        return path
    sess_out = os.getenv("SESSION_OUTPUT_DIR", "")
    if sess_out:
        alt = os.path.join(sess_out, "dataset", os.path.basename(path))
        if os.path.exists(alt):
            return alt
    return path


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", default=None, help="需求分析 Excel 文件路径")
    parser.add_argument("--output", default=None, help="输出文件路径")
    parser.add_argument("--per-scenario", type=int, default=2, help="每个场景生成几条用例（默认2）")
    parser.add_argument("--dimensions", default="", help="指定维度 ID 列表，逗号分隔（默认全部）")
    parser.add_argument("--list", action="store_true", help="列出所有可用维度并退出")
    args = parser.parse_args()

    script_dir = os.path.dirname(os.path.abspath(__file__))
    data_dir = os.path.join(script_dir, "..", "data")

    input_path = args.input
    if not input_path:
        input_path = os.path.join(data_dir, "requirements_analysis.xlsx")
    input_path = _resolve_path(input_path)

    if not os.path.exists(input_path):
        if args.list:
            print(f"[ERROR] 未找到需求分析文件", file=sys.stderr)
            print(f"      尝试路径: {input_path}", file=sys.stderr)
            return
        else:
            print(f"[ERROR] 输入文件不存在: {input_path}", file=sys.stderr)
            sys.exit(1)

    data_output = os.path.join(data_dir, "test_cases.xlsx")

    output_path = args.output
    if not output_path:
        output_path = data_output
    output_path = os.path.abspath(output_path)

    dims = _read_dimensions(input_path)

    if args.list:
        print(f"可用维度 ({len(dims)} 个):")
        for d in dims:
            print(f"  {d['id']} - {d['name']} ({d['type']})")
        print()
        print("请回复：每个场景生成几条用例？全部维度还是指定维度？")
        print("（指定维度用 --dimensions DIM-001,DIM-002 参数）")
        return

    api_key = os.getenv("LLM_API_KEY", "")
    model = os.getenv("LLM_MODEL", "gpt-4o")
    base_url = os.getenv("LLM_BASE_URL", "https://api.openai.com/v1")
    timeout = int(os.getenv("LLM_TIMEOUT", "300"))

    if not api_key:
        print("[ERROR] LLM_API_KEY not set", file=sys.stderr)
        sys.exit(1)

    scenarios = _read_scenarios(input_path)
    if not scenarios:
        print("[ERROR] 未找到测试场景", file=sys.stderr)
        sys.exit(1)

    # 按维度过滤
    selected_dims = [d.strip() for d in args.dimensions.split(",") if d.strip()] if args.dimensions else []
    if selected_dims:
        before = len(scenarios)
        scenarios = [s for s in scenarios if s["dimension_id"] in selected_dims]
        print(f"dimension filter: {before} -> {len(scenarios)} scenarios", file=sys.stderr)

    print(f"scenarios count: {len(scenarios)}, per_scenario: {args.per_scenario}", file=sys.stderr)

    test_cases = _generate_all(api_key, model, base_url, scenarios, args.per_scenario, timeout)

    _write_excel(test_cases, output_path)
    if output_path != data_output:
        _write_excel(test_cases, data_output)
    print(f"test cases count: {len(test_cases)}", file=sys.stderr)
    print("[阶段 2/4] 测试用例生成完成", file=sys.stderr)
    print(f"产出文件：{output_path}", file=sys.stderr)
    print("[阶段 2/4] 测试用例生成完成")
    print("产出文件：data/test_cases.xlsx")


def _read_dimensions(input_path):
    from openpyxl import load_workbook
    wb = load_workbook(input_path, read_only=True)
    if "测试维度" not in wb.sheetnames:
        return []
    ws = wb["测试维度"]
    dims = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        dims.append({"id": str(row[0] or ""), "name": str(row[1] or ""), "type": str(row[2] or "")})
    return dims


def _read_scenarios(input_path):
    from openpyxl import load_workbook
    wb = load_workbook(input_path, read_only=True)
    if "测试场景" not in wb.sheetnames:
        return []
    ws = wb["测试场景"]
    scenarios = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if not row or not row[0]:
            continue
        scenarios.append({
            "id": str(row[0] or ""),
            "dimension_id": str(row[1] or ""),
            "name": str(row[2] or ""),
            "description": str(row[3] or ""),
        })
    return scenarios


def _generate_all(api_key, model, base_url, scenarios, per_scenario, timeout):
    """分批调用 LLM 生成测试用例"""
    all_cases = []
    batch_size = 10
    for i in range(0, len(scenarios), batch_size):
        batch = scenarios[i:i + batch_size]
        cases = _call_llm_batch(api_key, model, base_url, batch, per_scenario, timeout)
        if cases:
            all_cases.extend(cases)
        print(f"batch {i // batch_size + 1}/{(len(scenarios) - 1) // batch_size + 1} done", file=sys.stderr)
    return all_cases


def _call_llm_batch(api_key, model, base_url, batch, per_scenario, timeout):
    system_prompt = (
        "# Role: 智能体（Agent）高级评测工程师,也是一位资深的高级软件测试工程师"
        "## Profile: 你专注于 Agent 系统的质量保障。你熟练掌握 ReAct、CoT（思维链）等 Agent 交互模式。你设计的用例旨在验证 Agent 在动态环境下的决策准确性、工具调用率以及任务最终达成率，也精通边界值分析法、等价类划分法、状态迁移法及各类黑盒测试技巧。你不仅能设计常规功能用例，还能设计深度的异常、边界和安全用例。请你根据测试场景生成详细可执行的测试用例。\n\n"
        "## Task: 请根据我提供的测试维度和测试场景，设计一套针对 Agent智能体功能 的测试用例。"
        "## Requirements: 每个测试用例包含：\n"
        "- 用例 ID（TC-{NNNN}）\n"
        "- 场景引用（SC-{XXX}）\n"
        "- 标题（简短描述）\n"
        "- 优先级（高/中/低）\n"
        "- 前置条件（执行前的系统状态）\n"
        "- 测试步骤（编号列表）\n"
        "- 用户输入\n"
        "- 预期结果（可验证的断言）\n\n"
        "设计原则：原子性（每个用例验证一个行为）、确定性（步骤无歧义）、自包含（数据内联）、可追溯性（链接回场景ID）\n\n"
        "请严格按照 JSON 格式输出，不要包含 markdown 标记。"
    )
    user_prompt = f"""请为以下 {len(batch)} 个测试场景生成测试用例。每个场景生成 {per_scenario} 个用例。

场景列表：
{json.dumps(batch, ensure_ascii=False, indent=2)}

请按以下 JSON 格式输出（只输出 JSON，不要 markdown 标记）：
{{
  "test_cases": [
    {{
      "scenario_id": "SC-001",
      "tc_id": "TC-0001",
      "title": "标题",
      "priority": "高",
      "preconditions": "前置条件",
      "steps": ["步骤1", "步骤2"],
      "user_input": "用户输入的文本",
      "expected": "预期结果"
    }}
  ]
}}
"""
    url = base_url.rstrip("/") + "/chat/completions"
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
    }
    body = {
        "model": model,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ],
        "temperature": 0.3,
    }
    extra_opts = os.getenv("LLM_EXTRA_OPTS", "")
    if extra_opts:
        try:
            body.update(json.loads(extra_opts))
        except json.JSONDecodeError:
            pass

    try:
        resp = requests.post(url, headers=headers, json=body, timeout=timeout)
        resp.raise_for_status()
        content = resp.json()["choices"][0]["message"]["content"]
        data = _extract_json(content)
        return data.get("test_cases", [])
    except Exception as e:
        print(f"[WARN] batch LLM call failed: {e}", file=sys.stderr)
        return []


def _extract_json(raw: str):
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        pass
    m = re.search(r"```(?:json)?\s*([\s\S]*?)```", raw)
    if m:
        try:
            return json.loads(m.group(1))
        except json.JSONDecodeError:
            pass
    m = re.search(r"\{[\s\S]*\}", raw)
    if m:
        try:
            return json.loads(m.group(0))
        except json.JSONDecodeError:
            pass
    return {}


def _write_excel(test_cases: list, path: str):
    if os.path.exists(path):
        os.remove(path)
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"
    ws.append(["用例 ID", "场景引用", "标题", "优先级", "前置条件", "测试步骤", "用户输入", "预期结果", "状态"])
    for tc in test_cases:
        steps = "\n".join(tc.get("steps", []))
        ws.append([
            tc.get("tc_id", ""),
            tc.get("scenario_id", ""),
            tc.get("title", ""),
            tc.get("priority", ""),
            tc.get("preconditions", ""),
            steps,
            tc.get("user_input", ""),
            tc.get("expected", ""),
            "未执行",
        ])
    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        import traceback
        traceback.print_exc()
        sys.exit(1)
