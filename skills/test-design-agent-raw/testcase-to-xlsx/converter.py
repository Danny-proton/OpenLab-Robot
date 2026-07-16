#!/usr/bin/env python3
"""
测试用例 Markdown 转 Excel 转换工具
将测试用例 markdown 文件 (.md) 转换为 Excel 格式 (.xlsx)
"""

import os
import re
import argparse
from pathlib import Path
import sys

try:
    import pandas as pd
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print(f"缺少依赖库：{e}")
    print("请先安装依赖：pip install pandas openpyxl")
    sys.exit(1)


def get_excel_path_from_md(md_path):
    """根据 markdown 文件路径生成对应的 excel 文件路径"""
    return os.path.splitext(md_path)[0] + '.xlsx'


def parse_testcase_md(content):
    """解析测试用例 markdown 内容"""
    cases = {
        '功能测试用例': [],
        '场景测试用例': [],
        '场景化 DFX 测试用例': []
    }

    # 移除标题和统计表格，只保留用例内容
    content = re.sub(r'^#.*?\n', '', content)
    content = re.sub(r'^##.*?\n', '', content)
    content = re.sub(r'^\|.*?\n', '', content, flags=re.MULTILINE)
    content = re.sub(r'^\+-+.*?\n', '', content, flags=re.MULTILINE)

    # 分割用例：先按 --- 分割，再提取用例
    raw_cases = re.split(r'\n---\n+', content)

    for raw_case in raw_cases:
        raw_case = raw_case.strip()
        if not raw_case or '【用例编号】' not in raw_case:
            continue

        case_data = {}

        # 提取用例编号
        num_match = re.search(r'【用例编号】\s*\n*([^\n]+)', raw_case)
        if num_match:
            case_data['用例编号'] = num_match.group(1).strip()
        else:
            continue

        # 提取用例名称
        name_match = re.search(r'【用例名称】\s*\n*([^\n]+)', raw_case)
        if name_match:
            case_data['用例名称'] = name_match.group(1).strip()
        else:
            continue

        # 提取预置条件
        cond_match = re.search(r'【预置条件】\s*\n*([\s\S]*?)(?=【测试步骤】|【预期结果】|【设计描述】|\n###|\Z)', raw_case)
        if cond_match:
            case_data['预置条件'] = cond_match.group(1).strip()
        else:
            case_data['预置条件'] = ''

        # 提取测试步骤
        steps_match = re.search(r'【测试步骤】\s*\n*([\s\S]*?)(?=【预期结果】|【设计描述】|\n###|\Z)', raw_case)
        if steps_match:
            case_data['测试步骤'] = steps_match.group(1).strip()
        else:
            case_data['测试步骤'] = ''

        # 提取预期结果
        result_match = re.search(r'【预期结果】\s*\n*([\s\S]*?)(?=【设计描述】|\n###|\Z)', raw_case)
        if result_match:
            case_data['预期结果'] = result_match.group(1).strip()
        else:
            case_data['预期结果'] = ''

        # 提取设计描述
        desc_match = re.search(r'【设计描述】\s*\n*([\s\S]*?)(?=\n\n|\n【用例编号】|\Z)', raw_case)
        if desc_match:
            case_data['设计描述'] = desc_match.group(1).strip()
        else:
            case_data['设计描述'] = ''

        # 根据用例编号前缀分类
        case_num = case_data['用例编号']
        if case_num.startswith('FUN_'):
            cases['功能测试用例'].append(case_data)
        elif case_num.startswith('SCN_'):
            cases['场景测试用例'].append(case_data)
        elif case_num.startswith('DFX_'):
            cases['场景化 DFX 测试用例'].append(case_data)

    return cases


def clean_text(text):
    """清理文本，移除多余空白"""
    if not text:
        return ''
    # 保留换行符但清理前后空白
    lines = text.split('\n')
    cleaned_lines = [line.strip() for line in lines if line.strip()]
    return '\n'.join(cleaned_lines)


def create_excel(cases, output_path):
    """创建 Excel 文件"""
    wb = Workbook()

    # 定义样式
    bold_font = Font(bold=True, size=11)
    normal_font = Font(size=10)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

    def add_sheet(sheet_name, data):
        """添加数据到工作表"""
        ws = wb.create_sheet(title=sheet_name[:31])  # Excel 工作表名最多 31 个字符

        if not data:
            ws['A1'] = f'{sheet_name} - 无数据'
            return

        headers = ['用例编号', '用例名称', '预置条件', '测试步骤', '预期结果', '设计描述']

        # 写入表头
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = bold_font
            cell.alignment = center_align
            cell.fill = header_fill
            cell.border = thin_border

        # 写入数据
        for row_idx, case in enumerate(data, 2):
            ws.cell(row=row_idx, column=1, value=case['用例编号']).border = thin_border
            ws.cell(row=row_idx, column=2, value=case['用例名称']).border = thin_border
            ws.cell(row=row_idx, column=3, value=clean_text(case['预置条件'])).alignment = left_align
            ws.cell(row=row_idx, column=3).border = thin_border
            ws.cell(row=row_idx, column=4, value=clean_text(case['测试步骤'])).alignment = left_align
            ws.cell(row=row_idx, column=4).border = thin_border
            ws.cell(row=row_idx, column=5, value=clean_text(case['预期结果'])).alignment = left_align
            ws.cell(row=row_idx, column=5).border = thin_border
            ws.cell(row=row_idx, column=6, value=clean_text(case['设计描述'])).alignment = left_align
            ws.cell(row=row_idx, column=6).border = thin_border

        # 调整列宽
        column_widths = {
            1: 25,  # 用例编号
            2: 30,  # 用例名称
            3: 40,  # 预置条件
            4: 50,  # 测试步骤
            5: 50,  # 预期结果
            6: 60   # 设计描述
        }

        for col, width in column_widths.items():
            ws.column_dimensions[get_column_letter(col)].width = width

        # 设置行高
        ws.row_dimensions[1].height = 25
        for row in range(2, len(data) + 2):
            ws.row_dimensions[row].height = 20

    # 添加各 Sheet
    for sheet_name, data in cases.items():
        add_sheet(sheet_name, data)

    # 删除默认 Sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # 保存文件
    wb.save(output_path)
    print(f"Excel 文件已保存至：{output_path}")


def convert_markdown_to_excel(input_path, output_path=None):
    """
    适配 Claude 技能调用的接口，通过环境变量或参数传递路径。

    参数:
        input_path: Markdown 文件路径
        output_path: Excel 文件路径（可选，默认：与输入文件同目录）

    返回:
        bool: 转换是否成功
    """
    # 如果未指定输出路径，自动生成
    if output_path is None:
        output_path = get_excel_path_from_md(input_path)

    # 检查输入文件是否存在
    if not os.path.exists(input_path):
        print(f"错误：输入文件不存在：{input_path}")
        return False

    # 确保输出目录存在
    output_dir = os.path.dirname(output_path)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)
        print(f"已创建输出目录：{output_dir}")

    print(f"正在读取：{input_path}")

    # 读取 markdown 文件
    try:
        with open(input_path, 'r', encoding='utf-8') as f:
            content = f.read()
    except Exception as e:
        print(f"读取文件失败：{e}")
        return False

    # 解析测试用例
    print("正在解析测试用例...")
    cases = parse_testcase_md(content)

    # 统计
    total_cases = sum(len(c) for c in cases.values())
    print(f"共解析到 {total_cases} 个测试用例:")
    for sheet_name, data in cases.items():
        print(f"  - {sheet_name}: {len(data)} 个")

    # 创建 Excel 文件
    print("正在生成 Excel 文件...")
    create_excel(cases, output_path)

    print("转换完成!")
    return True


def main():
    parser = argparse.ArgumentParser(description='测试用例 Markdown 转 Excel 工具')
    parser.add_argument('--input', '-i', type=str, default=None,
                        help='输入 markdown 文件路径（或通过环境变量 INPUT_PATH 指定）')
    parser.add_argument('--output', '-o', type=str, default=None,
                        help='输出 Excel 文件路径（或通过环境变量 OUTPUT_PATH 指定）')

    args = parser.parse_args()

    # 优先级：命令行参数 > 环境变量
    input_path = args.input or os.environ.get('INPUT_PATH')
    output_path = args.output or os.environ.get('OUTPUT_PATH')

    if not input_path:
        print("错误：必须指定输入文件路径（--input 或 INPUT_PATH 环境变量）")
        sys.exit(1)

    # 自动生成输出路径（与输入文件同目录）
    if output_path is None:
        output_path = get_excel_path_from_md(input_path)

    # 确保输出目录存在
    output_dir = os.path.dirname(output_path)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)
        print(f"已创建输出目录：{output_dir}")

    # 检查输入文件
    if not os.path.exists(input_path):
        print(f"错误：输入文件不存在：{input_path}")
        sys.exit(1)

    print(f"正在读取：{input_path}")

    # 读取 markdown 文件
    try:
        with open(input_path, 'r', encoding='utf-8') as f:
            content = f.read()
    except Exception as e:
        print(f"读取文件失败：{e}")
        sys.exit(1)

    # 解析测试用例
    print("正在解析测试用例...")
    cases = parse_testcase_md(content)

    # 统计
    total_cases = sum(len(c) for c in cases.values())
    print(f"共解析到 {total_cases} 个测试用例:")
    for sheet_name, data in cases.items():
        print(f"  - {sheet_name}: {len(data)} 个")

    # 创建 Excel 文件
    print("正在生成 Excel 文件...")
    create_excel(cases, output_path)

    print("转换完成!")

    # Windows 环境下自动打开文件
    if os.name == 'nt':
        try:
            os.startfile(output_path)
            print(f"已打开文件：{output_path}")
        except:
            pass


if __name__ == '__main__':
    main()
