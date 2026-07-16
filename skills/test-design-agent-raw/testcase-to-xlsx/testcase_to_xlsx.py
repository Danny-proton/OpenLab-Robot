#!/usr/bin/env python3
"""
测试用例 Markdown 转 Excel Skill - 支持多种用例格式的通用解析脚本

此脚本支持多种测试用例格式：
1. 标准格式：用例编号、用例名称、预置条件、测试步骤、预期结果、设计描述
2. 带分隔线的格式：用例之间使用 ### 标题分隔
3. 连续格式：多个用例连续排列，通过用例编号分隔
4. 两种字段标记：**字段名**：和【字段名】
"""

import os
import re
import argparse
import sys
import subprocess
from pathlib import Path


def detect_python_executable():
    """自动检测合适的 Python 可执行文件"""
    system_root = os.environ.get('SystemRoot', 'C:\\Windows')
    python_paths = [
        f"{system_root}\\System32\\python.exe",
        f"{system_root}\\SysWOW64\\python.exe",
        "python",
        "python3",
    ]

    for path in python_paths:
        try:
            result = subprocess.run([path, '--version'],
                                  capture_output=True,
                                  text=True,
                                  timeout=5)
            if result.returncode == 0:
                return path
        except:
            continue

    common_paths = [
        "C:\\Python310\\python.exe",
        "C:\\Python39\\python.exe",
        "C:\\Python38\\python.exe",
        "C:\\Users\\l30064969\\AppData\\Local\\Programs\\Python\\Python310\\python.exe",
    ]

    for path in common_paths:
        if os.path.exists(path):
            return path

    return "python"


def run_with_correct_python(script_path, args):
    """使用正确的 Python 解释器运行脚本"""
    python_exe = detect_python_executable()

    if os.name == 'nt':
        cmd = [python_exe, str(script_path)] + list(args)
    else:
        cmd = [python_exe, str(script_path)] + list(args)

    try:
        result = subprocess.run(cmd, capture_output=False, text=True)
        return result.returncode
    except Exception as e:
        print(f"运行脚本失败：{e}")
        return 1


def get_excel_path_from_md(md_path):
    """根据 markdown 文件路径生成对应的 excel 文件路径"""
    return os.path.splitext(md_path)[0] + '.xlsx'


def split_cases_by_pattern(raw_content):
    """
    使用正则分割用例，兼容多种格式
    支持：
    1. ### 标题分隔的格式（用例编号在 ### 标题后）
    2. 空行分隔的格式
    3. 连续用例通过编号分隔的格式（直接按用例编号分割）
    """
    # 统一使用半角冒号模式
    case_num_pattern_str = r'(\*\*用例编号\*\*|【用例编号】)\s*[:：]'
    case_num_pattern = re.compile(case_num_pattern_str, re.MULTILINE)

    # 直接使用模式 3：按用例编号分割
    # 这是最可靠的方式，适用于所有用例格式
    cases = []
    matches = list(case_num_pattern.finditer(raw_content))

    for i, match in enumerate(matches):
        start = match.start()
        if i + 1 < len(matches):
            end = matches[i + 1].start()
        else:
            end = len(raw_content)

        case_content = raw_content[start:end].strip()
        if case_content:
            cases.append(case_content)

    return cases


def extract_field_value(content, field_name):
    """
    提取字段值，支持两种格式：
    1. **字段名**：值
    2. **字段名**:\n值
    """
    # 支持两种字段标记和两种冒号（半角：和全角：）
    patterns = [
        # **字段名**：换行后值（支持中文冒号）
        rf'\*\*{field_name}\*\*\s*[:：]\s*\n*\s*([^\n]+)',
        # **字段名**：值在同一行
        rf'\*\*{field_name}\*\*\s*[:：]\s*([^\n]+)',
        # 【字段名】：换行后值
        rf'【{field_name}】\s*[:：]\s*\n*\s*([^\n]+)',
        # 【字段名】：值在同一行
        rf'【{field_name}】\s*[:：]\s*([^\n]+)',
    ]

    for pattern in patterns:
        match = re.search(pattern, content)
        if match:
            return match.group(1).strip()

    return ''


def extract_field_list_value(content, field_name, max_matches=None):
    """
    提取列表类型的字段值（预置条件、测试步骤、预期结果）
    返回多行文本
    """
    # 使用更精确的模式：匹配当前字段的内容
    # 找到当前字段的位置，然后找到下一个字段或文档结尾
    field_start_pattern = rf'\*\*{field_name}\*\*\s*[:：]'
    field_start_match = re.search(field_start_pattern, content)

    if not field_start_match:
        return ''

    start_pos = field_start_match.end()

    # 定义可能的终止符（支持中文和英文冒号）
    # 使用 re.escape 转义特殊字符
    next_fields_raw = ['**用例编号**:', '**用例名称**:', '**预置条件**:',
                       '**测试步骤**:', '**预期结果**:', '**设计描述**:',
                       '【用例编号】:', '【用例名称】:', '【预置条件】:',
                       '【测试步骤】:', '【预期结果】:', '【设计描述】:']
    # 也支持中文冒号
    next_fields_zh_raw = [f[:-1] + '：' for f in next_fields_raw]
    next_fields_all = [re.escape(f) for f in next_fields_raw + next_fields_zh_raw]

    next_headers = ['\n##', '\n###']

    # 查找下一个字段或标题的位置
    end_pos = len(content)

    # 查找下一个字段的位置
    for next_field in next_fields_all:
        try:
            next_match = re.search(next_field, content[start_pos:])
            if next_match:
                next_pos = start_pos + next_match.start()
                if next_pos < end_pos:
                    end_pos = next_pos
        except re.error:
            continue

    # 查找下一个标题的位置
    for header in next_headers:
        header_match = re.search(header, content[start_pos:])
        if header_match:
            header_pos = start_pos + header_match.start()
            if header_pos < end_pos:
                end_pos = header_pos

    # 提取字段内容
    field_content = content[start_pos:end_pos].strip()

    # 清理内容：移除空白行但保留序号，同时清理行首的列表符号（如"-"、"-"等）
    lines = []
    for line in field_content.split('\n'):
        line = line.strip()
        if not line:
            continue
        # 清理行首的列表符号（如 "- "、"  - " 等）
        line = re.sub(r'^[\s\-\••]\s*', '', line)
        # 清理行首的缩进数字列表（保留数字列表格式）
        line = line.lstrip()
        if line:
            lines.append(line)

    return '\n'.join(lines)


def parse_testcase_md(content):
    """解析测试用例 markdown 内容，支持多种格式"""
    cases = {
        '功能测试用例': [],
        '场景测试用例': [],
        'DFX 测试用例': []  # DFX 类型用例合并到一个 sheet
    }

    # 前缀到 Sheet 名称的映射
    prefix_to_sheet = {
        'FUN_': '功能测试用例',
        'SCN_': '场景测试用例',
    }

    # DFX 前缀列表
    dfx_prefixes = ['DFP_', 'DFR_', 'DFS_', 'DFC_', 'DFAI_', 'DFINT_']

    # 直接按用例编号分割，不修改原始内容
    raw_cases = split_cases_by_pattern(content)

    for idx, raw_case in enumerate(raw_cases):
        raw_case = raw_case.strip()
        if not raw_case:
            continue

        # 提取用例编号
        case_num = extract_field_value(raw_case, '用例编号')
        if not case_num:
            continue

        # 提取用例名称
        case_name = extract_field_value(raw_case, '用例名称')
        if not case_name:
            continue

        # 提取预置条件（列表格式）
        pre_cond = extract_field_list_value(raw_case, '预置条件')

        # 提取测试步骤（列表格式）
        steps = extract_field_list_value(raw_case, '测试步骤')

        # 提取预期结果（列表格式）
        results = extract_field_list_value(raw_case, '预期结果')

        # 提取设计描述（单行格式）
        desc = extract_field_value(raw_case, '设计描述')

        # 组装用例数据
        case_data = {
            '用例编号': case_num,
            '用例名称': case_name,
            '预置条件': pre_cond,
            '测试步骤': steps,
            '预期结果': results,
            '设计描述': desc
        }

        # 根据用例编号前缀分类
        if case_num.startswith('FUN_'):
            cases['功能测试用例'].append(case_data)
        elif case_num.startswith('SCN_'):
            cases['场景测试用例'].append(case_data)
        else:
            # 所有 DFX 前缀的用例合并到 DFX 测试用例 sheet
            for prefix in dfx_prefixes:
                if case_num.startswith(prefix):
                    cases['DFX 测试用例'].append(case_data)
                    break

    return cases


def clean_text(text):
    """清理文本，移除多余空白但保留换行符"""
    if not text:
        return ''
    lines = text.split('\n')
    cleaned_lines = [line.strip() for line in lines if line.strip()]
    return '\n'.join(cleaned_lines)


def create_excel(cases, output_path):
    """创建 Excel 文件"""
    try:
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        print(f"缺少依赖库：{e}")
        print("请先安装依赖：pip install pandas openpyxl")
        sys.exit(1)

    wb = Workbook()

    # 定义样式
    bold_font = Font(bold=True, size=11)
    normal_font = Font(size=10)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

    def add_sheet(sheet_name, data):
        """添加数据到工作表"""
        ws = wb.create_sheet(title=sheet_name[:31])

        if not data:
            ws['A1'] = f'{sheet_name} - 无数据'
            return

        headers = ['用例编号', '用例名称', '预置条件', '测试步骤', '预期结果', '设计描述']

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = bold_font
            cell.alignment = center_align
            cell.fill = header_fill
            cell.border = thin_border

        for row_idx, case in enumerate(data, 2):
            ws.cell(row=row_idx, column=1, value=case['用例编号']).border = thin_border
            ws.cell(row=row_idx, column=2, value=case['用例名称']).border = thin_border
            ws.cell(row=row_idx, column=3, value=clean_text(case['预置条件'])).alignment = left_align
            ws.cell(row=row_idx, column=3).border = thin_border
            ws.cell(row=row_idx, column=4, value=clean_text(case['测试步骤'])).alignment = left_align
            ws.cell(row=row_idx, column=4).border = thin_border
            ws.cell(row=row_idx, column=5, value=clean_text(case['预期结果'])).alignment = left_align
            ws.cell(row=row_idx, column=5).border = thin_border
            ws.cell(row=row_idx, column=6, value=clean_text(case['设计描述'])).alignment = left_align
            ws.cell(row=row_idx, column=6).border = thin_border

        # 调整列宽
        column_widths = {
            1: 25,  # 用例编号
            2: 35,  # 用例名称
            3: 45,  # 预置条件
            4: 50,  # 测试步骤
            5: 50,  # 预期结果
            6: 60   # 设计描述
        }

        for col, width in column_widths.items():
            ws.column_dimensions[get_column_letter(col)].width = width

        # 设置行高
        ws.row_dimensions[1].height = 25
        for row in range(2, len(data) + 2):
            ws.row_dimensions[row].height = 20

    # 添加各 Sheet
    for sheet_name, data in cases.items():
        add_sheet(sheet_name, data)

    # 删除默认 Sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # 保存文件
    wb.save(output_path)
    print(f"Excel 文件已保存至：{output_path}")


def main():
    """主入口"""
    parser = argparse.ArgumentParser(description='测试用例 Markdown 转 Excel 工具')
    parser.add_argument('--input', '-i', type=str, default=None,
                        help='输入 markdown 文件路径')
    parser.add_argument('--output', '-o', type=str, default=None,
                        help='输出 Excel 文件路径')

    args = parser.parse_args()

    # 优先级：命令行参数 > 环境变量
    input_path = args.input or os.environ.get('INPUT_PATH')
    output_path = args.output or os.environ.get('OUTPUT_PATH')

    if not input_path:
        print("错误：必须指定输入文件路径（--input 或 INPUT_PATH 环境变量）")
        sys.exit(1)

    # 自动生成输出路径
    if output_path is None:
        output_path = get_excel_path_from_md(input_path)

    # 确保输出目录存在
    output_dir = os.path.dirname(output_path)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # 检查输入文件
    if not os.path.exists(input_path):
        print(f"错误：输入文件不存在：{input_path}")
        sys.exit(1)

    # 读取 markdown 文件
    try:
        with open(input_path, 'r', encoding='utf-8') as f:
            content = f.read()
    except Exception as e:
        print(f"读取文件失败：{e}")
        sys.exit(1)

    # 解析测试用例
    cases = parse_testcase_md(content)

    # 统计
    total_cases = sum(len(c) for c in cases.values())

    # 创建 Excel 文件
    create_excel(cases, output_path)

    # Windows 环境下自动打开文件
    if os.name == 'nt':
        try:
            os.startfile(output_path)
            print(f"已打开文件：{output_path}")
        except:
            pass


if __name__ == '__main__':
    main()
