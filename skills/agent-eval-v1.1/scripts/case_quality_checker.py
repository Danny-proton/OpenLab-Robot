#!/usr/bin/env python3
"""case_quality_checker.py — 12 维度用例质量检查（确定性，零 LLM）。

V1.1 用例自优化的分析模块之一。对 cases YAML 做 12 维质量评分：
- 9 维（继承 test-design-agent-raw + 业界标准）
- 3 维 Agent 专属（工具覆盖率/工作流覆盖率/记忆覆盖率）

权重合计 1.00。阈值：单维 < 0.6 标记低分；总分 < 0.75 触发质量增强。

用法:
  python case_quality_checker.py --config .agent-eval/config.yaml --split train
  python case_quality_checker.py --config .agent-eval/config.yaml --split train --out quality.json
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

sys.path.insert(0, str(Path(__file__).resolve().parent))
import common as C  # noqa: E402
import case_io as CIO  # noqa: E402


# ---------------------------------------------------------------------------
# 12 维度定义 + 权重
# ---------------------------------------------------------------------------

DIMENSIONS = [
    # (id, name, weight, is_agent_specific)
    ("spec_completeness",   "SPEC 完整性",   0.13, False),
    ("case_completeness",   "用例完整性",    0.09, False),
    ("feature_coverage",    "功能点覆盖度",  0.13, False),
    ("dfx_coverage",        "DFX 覆盖度",    0.13, False),
    ("valid_case_ratio",    "有效用例率",    0.09, False),
    ("executability",       "执行可行性",    0.09, False),
    ("ambiguity_free",      "无二义性",      0.09, False),
    ("length_reasonable",   "长度合理",      0.04, False),
    ("assertion_verifiable","断言可验证",    0.09, False),
    ("tool_coverage",       "工具覆盖率",    0.04, True),
    ("workflow_coverage",   "工作流覆盖率",  0.04, True),
    ("memory_coverage",     "记忆覆盖率",    0.04, True),
]

# DFX 7 维
DFX_TYPES = ["dfx_security", "dfx_reliability", "dfx_performance",
             "dfx_compatibility", "dfx_serviceability", "dfx_maintainability"]

# 二义性关键词（启发式）
AMBIGUITY_WORDS = ["等等", "之类的", "可能", "大概", "也许", "或者", "之类", "等等"]

# 10 功能维度（若 requirements Excel 不存在，用 case 的 dimension_id 推断）
DEFAULT_FEATURE_DIMS = 10

LOW_SCORE_THRESHOLD = 0.6
TOTAL_THRESHOLD = 0.75


# ---------------------------------------------------------------------------
# 单维度检查
# ---------------------------------------------------------------------------

def check_spec_completeness(cases: list[dict], req_dimensions: list[str] | None) -> tuple[float, dict]:
    """维度1: 用例覆盖需求维度的比例。"""
    if req_dimensions:
        total = len(req_dimensions)
        covered = set()
        for c in cases:
            d = c.get("dimension_id")
            if d:
                covered.add(d)
        score = len(covered) / total if total > 0 else 0.0
    else:
        # 无 requirements，用 dimension_id 去重数 / DEFAULT_FEATURE_DIMS
        covered = set(c.get("dimension_id") for c in cases if c.get("dimension_id"))
        total = DEFAULT_FEATURE_DIMS
        score = min(len(covered) / total, 1.0) if total > 0 else 0.0
    return score, {"covered": len(covered), "total": total}


def check_case_completeness(cases: list[dict]) -> tuple[float, dict]:
    """维度2: 字段齐全率。"""
    required_fields = ["id", "name", "agent", "task", "input", "expected",
                       "expected_tools", "business_rules", "expected_steps", "scoring"]
    if not cases:
        return 0.0, {"total": 0, "complete": 0}
    complete = 0
    missing_fields: dict[str, list[str]] = {}
    for c in cases:
        miss = [f for f in required_fields if f not in c or c[f] is None]
        if not miss:
            complete += 1
        else:
            missing_fields[c.get("id", "?")] = miss
    score = complete / len(cases)
    return score, {"complete": complete, "total": len(cases), "missing": missing_fields}


def check_feature_coverage(cases: list[dict], req_dimensions: list[str] | None) -> tuple[float, dict]:
    """维度3: 功能维度覆盖。"""
    covered = set(c.get("dimension_id") for c in cases if c.get("dimension_id"))
    if req_dimensions:
        total = len(req_dimensions)
        score = len(covered & set(req_dimensions)) / total if total > 0 else 0.0
    else:
        total = DEFAULT_FEATURE_DIMS
        score = min(len(covered) / total, 1.0)
    return score, {"covered_dims": sorted(covered), "total": total}


def check_dfx_coverage(cases: list[dict]) -> tuple[float, dict]:
    """维度4: DFX 7 维覆盖。"""
    covered = set()
    for c in cases:
        cat = c.get("category", "functional")
        if cat in DFX_TYPES:
            covered.add(cat)
    total = len(DFX_TYPES)
    score = len(covered) / total
    return score, {"covered_dfx": sorted(covered), "total": total,
                   "missing": sorted(set(DFX_TYPES) - covered)}


def check_valid_case_ratio(cases: list[dict]) -> tuple[float, dict]:
    """维度5: 有效用例率（lifecycle != deprecated）。"""
    if not cases:
        return 0.0, {"total": 0, "valid": 0}
    valid = [c for c in cases if c.get("lifecycle", "active") != "deprecated"]
    score = len(valid) / len(cases)
    return score, {"valid": len(valid), "total": len(cases),
                   "deprecated": len(cases) - len(valid)}


def check_executability(cases: list[dict], scores: dict | None) -> tuple[float, dict]:
    """维度6: 执行可行性（上次执行通过率 > 0）。"""
    if not scores or not scores.get("per_case"):
        # 无历史分数，按用例字段完整性推断可执行性
        executable = 0
        for c in cases:
            inp = c.get("input") or {}
            if inp.get("user_message") or inp.get("application_id"):
                executable += 1
        score = executable / len(cases) if cases else 0.0
        return score, {"executable": executable, "total": len(cases), "source": "inferred"}
    per_case = scores["per_case"]
    passed = sum(1 for pc in per_case if pc.get("weighted_score", 0) > 0)
    total = len(per_case)
    score = passed / total if total > 0 else 0.0
    return score, {"passed": passed, "total": total, "source": "scores"}


def check_ambiguity_free(cases: list[dict]) -> tuple[float, dict]:
    """维度7: 无二义性（启发式：含模糊词算二义）。"""
    if not cases:
        return 0.0, {"ambiguous": 0, "total": 0}
    ambiguous_ids: list[str] = []
    for c in cases:
        text = json.dumps(c, ensure_ascii=False).lower()
        if any(w in text for w in AMBIGUITY_WORDS):
            ambiguous_ids.append(c.get("id", "?"))
    score = 1 - len(ambiguous_ids) / len(cases)
    return score, {"ambiguous": len(ambiguous_ids), "total": len(cases),
                   "ambiguous_ids": ambiguous_ids}


def check_length_reasonable(cases: list[dict]) -> tuple[float, dict]:
    """维度8: 长度合理（expected_steps < 10，task token < 500）。"""
    if not cases:
        return 0.0, {"unreasonable": 0, "total": 0}
    unreasonable_ids: list[str] = []
    for c in cases:
        es = c.get("expected_steps", 8)
        task = c.get("task", "") or ""
        # 粗估 token：字符数 / 2（中文约 1.5 字/token，英文约 4 字符/token，取折中）
        task_tokens = len(task) / 2
        if es > 10 or task_tokens > 500:
            unreasonable_ids.append(c.get("id", "?"))
    score = 1 - len(unreasonable_ids) / len(cases)
    return score, {"unreasonable": len(unreasonable_ids), "total": len(cases),
                   "unreasonable_ids": unreasonable_ids}


def check_assertion_verifiable(cases: list[dict]) -> tuple[float, dict]:
    """维度9: 断言可验证（expected 非纯主观）。"""
    if not cases:
        return 0.0, {"verifiable": 0, "total": 0}
    verifiable = 0
    unverifiable_ids: list[str] = []
    for c in cases:
        exp = c.get("expected") or {}
        fd = exp.get("final_decision") or {}
        # 至少有一种可机器验证的断言
        if any(k in fd for k in ["contains", "equals", "regex", "schema"]):
            verifiable += 1
        else:
            unverifiable_ids.append(c.get("id", "?"))
    score = verifiable / len(cases)
    return score, {"verifiable": verifiable, "total": len(cases),
                   "unverifiable_ids": unverifiable_ids}


def check_tool_coverage(cases: list[dict], all_tools: list[str] | None) -> tuple[float, dict]:
    """维度10 (Agent专属): 工具覆盖率。"""
    if all_tools is None:
        all_tools = CIO.extract_all_tools(cases)
    if not all_tools:
        return 1.0, {"covered": 0, "total": 0, "note": "无工具定义"}
    covered: set[str] = set()
    for c in cases:
        et = c.get("expected_tools") or {}
        for t in et.get("required", []) or []:
            covered.add(t)
    score = len(covered & set(all_tools)) / len(all_tools)
    missing = sorted(set(all_tools) - covered)
    return score, {"covered": len(covered), "total": len(all_tools), "missing": missing}


def check_workflow_coverage(cases: list[dict]) -> tuple[float, dict]:
    """维度11 (Agent专属): 工作流覆盖率。

    检查 3 类工作流场景是否覆盖：
    - 前置检查（expect_preflight_advisor=true）
    - 异常恢复（category=dfx_reliability 或 task 含"异常/恢复/fallback"）
    - fallback（business_rules 含 fallback 相关）
    """
    workflow_types = {
        "preflight_check": False,
        "exception_recovery": False,
        "fallback": False,
    }
    for c in cases:
        if c.get("expect_preflight_advisor"):
            workflow_types["preflight_check"] = True
        cat = c.get("category", "")
        task = (c.get("task", "") or "").lower()
        if cat == "dfx_reliability" or any(w in task for w in ["异常", "恢复", "故障", "fallback"]):
            workflow_types["exception_recovery"] = True
        # fallback: 有 forbidden 工具说明有边界
        et = c.get("expected_tools") or {}
        if et.get("forbidden"):
            workflow_types["fallback"] = True
    covered = sum(1 for v in workflow_types.values() if v)
    total = len(workflow_types)
    score = covered / total
    return score, {"covered_types": [k for k, v in workflow_types.items() if v],
                   "missing_types": [k for k, v in workflow_types.items() if not v]}


def check_memory_coverage(cases: list[dict]) -> tuple[float, dict]:
    """维度12 (Agent专属): 记忆覆盖率。"""
    has_memory_case = any(c.get("expect_memory_use") for c in cases)
    # 也检查多轮场景（input 含 attachments 或 task 含"多轮/上下文/历史"）
    has_multi_turn = any(
        (c.get("input", {}).get("attachments")) or
        any(w in (c.get("task", "") or "") for w in ["多轮", "上下文", "历史", "之前"])
        for c in cases
    )
    covered = sum([has_memory_case, has_multi_turn])
    total = 2
    score = covered / total
    return score, {"has_memory_case": has_memory_case, "has_multi_turn": has_multi_turn}


# ---------------------------------------------------------------------------
# 主检查函数
# ---------------------------------------------------------------------------

def check_quality(
    cases: list[dict],
    req_dimensions: list[str] | None = None,
    scores: dict | None = None,
    all_tools: list[str] | None = None,
) -> dict[str, Any]:
    """跑 12 维质量检查，返回完整结果。"""
    if all_tools is None:
        all_tools = CIO.extract_all_tools(cases)

    checks = {
        "spec_completeness": lambda: check_spec_completeness(cases, req_dimensions),
        "case_completeness": lambda: check_case_completeness(cases),
        "feature_coverage": lambda: check_feature_coverage(cases, req_dimensions),
        "dfx_coverage": lambda: check_dfx_coverage(cases),
        "valid_case_ratio": lambda: check_valid_case_ratio(cases),
        "executability": lambda: check_executability(cases, scores),
        "ambiguity_free": lambda: check_ambiguity_free(cases),
        "length_reasonable": lambda: check_length_reasonable(cases),
        "assertion_verifiable": lambda: check_assertion_verifiable(cases),
        "tool_coverage": lambda: check_tool_coverage(cases, all_tools),
        "workflow_coverage": lambda: check_workflow_coverage(cases),
        "memory_coverage": lambda: check_memory_coverage(cases),
    }

    dim_results: dict[str, dict] = {}
    weighted_total = 0.0
    low_score_dims: list[str] = []

    for dim_id, dim_name, weight, is_agent in DIMENSIONS:
        score, detail = checks[dim_id]()
        dim_results[dim_id] = {
            "name": dim_name,
            "weight": weight,
            "score": round(score, 4),
            "agent_specific": is_agent,
            "detail": detail,
        }
        weighted_total += score * weight
        if score < LOW_SCORE_THRESHOLD:
            low_score_dims.append(dim_id)

    return {
        "dimensions": dim_results,
        "weighted_total": round(weighted_total, 4),
        "low_score_dimensions": low_score_dims,
        "total_threshold": TOTAL_THRESHOLD,
        "passes_threshold": weighted_total >= TOTAL_THRESHOLD,
        "n_cases": len(cases),
    }


# ---------------------------------------------------------------------------
# 辅助：从 requirements Excel 提取维度（可选）
# ---------------------------------------------------------------------------

def load_req_dimensions(cfg: C.EvalConfig) -> list[str] | None:
    """从 data/requirements_analysis.xlsx 提取维度 ID 列表（可选）。"""
    try:
        import openpyxl
    except ImportError:
        return None
    path = cfg.root.parent / "data" / "requirements_analysis.xlsx"
    if not path.exists():
        # 也尝试 skill 的 data 目录
        path = C.skill_dir() / "data" / "requirements_analysis.xlsx"
        if not path.exists():
            return None
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if "测试维度" not in wb.sheetnames:
            return None
        ws = wb["测试维度"]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return None
        header = rows[0]
        # 找维度 ID 列
        id_col = None
        for i, h in enumerate(header):
            if h and "维度" in str(h) and "id" in str(h).lower():
                id_col = i
                break
        if id_col is None:
            id_col = 0  # 默认第一列
        dims = []
        for row in rows[1:]:
            if row and row[id_col]:
                dims.append(str(row[id_col]))
        return dims if dims else None
    except Exception:
        return None


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main() -> int:
    ap = argparse.ArgumentParser(description="12 维用例质量检查")
    ap.add_argument("--config", required=True)
    ap.add_argument("--split", default="train")
    ap.add_argument("--run", help="关联的 run_id（用于读 scores 推断可执行性）")
    ap.add_argument("--out", help="输出 JSON 路径")
    args = ap.parse_args()

    cfg = C.EvalConfig.load(Path(args.config).resolve())
    cases = CIO.load_cases(cfg, args.split)

    req_dims = load_req_dimensions(cfg)

    scores = None
    if args.run:
        sp = cfg.scores_dir / f"{args.run}.json"
        if sp.exists():
            scores = json.loads(sp.read_text(encoding="utf-8"))

    result = check_quality(cases, req_dimensions=req_dims, scores=scores)

    # 输出
    out_path = Path(args.out) if args.out else (cfg.reports_dir / f"case_quality_{args.split}.json")
    C.write_json(out_path, result)

    # 打印摘要
    print(f"[case_quality_checker] split={args.split} cases={len(cases)}")
    print(f"[case_quality_checker] 加权总分: {result['weighted_total']:.4f} (阈值 {TOTAL_THRESHOLD})")
    print(f"[case_quality_checker] {'通过' if result['passes_threshold'] else '未通过'}阈值")
    print(f"[case_quality_checker] 低分维度（<{LOW_SCORE_THRESHOLD}）: {result['low_score_dimensions'] or '无'}")
    print(f"[case_quality_checker] 报告: {out_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
