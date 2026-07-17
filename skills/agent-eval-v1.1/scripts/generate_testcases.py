"""
手机银行智能体评测 - 测试用例生成（机械 I/O 层）

本脚本只做四件机械工作（不调用任何 LLM）：
  1. --list          : 列出需求分析 Excel 中的维度（供 Agent 决定生成范围）
  2. --read-scenarios: 把需求分析 Excel 中的场景读成 JSON 输出到 stdout（供 Agent 生成用例时引用）
  3. --write-stdin   : 把 Agent 生成好的用例 JSON 通过 stdin 写成 Excel（test_cases.xlsx）
  4. --write-file    : 同上，但从 --json-file 读

prompt 拼装和用例生成由 Agent（Claude）在子 skill
`skills/test-case-generator/SKILL.md` 中用文字指示，通过 Task 工具完成。
本脚本与任何外部模型 URL、API key 完全解耦。

Usage
-----
# 列出维度（决定生成范围）
python generate_testcases.py --list --input requirements_analysis.xlsx

# 输出场景 JSON 给 Agent 看
python generate_testcases.py --read-scenarios --input requirements_analysis.xlsx [--dimensions DIM-001,DIM-002]

# 写入用例（Agent 把生成的 JSON 通过 stdin 传入）
cat cases.json | python generate_testcases.py --write-stdin --input requirements_analysis.xlsx --output test_cases.xlsx

cases.json schema:
{
  "test_cases": [
    {
      "scenario_id": "SC-001",
      "tc_id": "TC-0001",
      "title": "标题",
      "priority": "高",
      "preconditions": "前置条件",
      "steps": ["步骤1", "步骤2"],
      "user_input": "用户输入",
      "expected": "预期结果",
      "assertion_type": "contains|exact|regex|schema|status_code"
    }
  ]
}
"""
import sys
import json
import os
import argparse


def main():
    parser = argparse.ArgumentParser(
        description="测试用例生成机械 I/O 层（无 LLM）。prompt 和生成由 Agent 在子 skill 中完成。"
    )
    parser.add_argument("--input", default=None, help="需求分析 Excel 文件路径")
    parser.add_argument("--output", default=None, help="输出用例 Excel 文件路径")
    parser.add_argument("--list", action="store_true", help="列出需求分析中的维度并退出")
    parser.add_argument("--read-scenarios", action="store_true",
                        help="把需求分析中的场景读成 JSON 输出到 stdout")
    parser.add_argument("--dimensions", default="",
                        help="过滤维度 ID，逗号分隔（配合 --read-scenarios）")
    parser.add_argument("--write-stdin", action="store_true",
                        help="从 stdin 读取 Agent 生成的用例 JSON，写成 Excel")
    parser.add_argument("--write-file", action="store_true",
                        help="从 --json-file 读取 Agent 生成的用例 JSON，写成 Excel")
    parser.add_argument("--json-file", default=None,
                        help="Agent 生成的用例 JSON 文件路径（配合 --write-file）")
    args = parser.parse_args()

    script_dir = os.path.dirname(os.path.abspath(__file__))
    data_dir = os.path.join(script_dir, "..", "data")

    input_path = args.input or os.path.join(data_dir, "requirements_analysis.xlsx")
    input_path = _resolve_path(input_path)

    # 模式 1: 列出维度
    if args.list:
        if not os.path.exists(input_path):
            print(f"[ERROR] 未找到需求分析文件: {input_path}", file=sys.stderr)
            sys.exit(1)
        dims = _read_dimensions(input_path)
        print(f"可用维度 ({len(dims)} 个):")
        for d in dims:
            print(f"  {d['id']} - {d['name']} ({d['type']})")
        return

    # 模式 2: 输出场景 JSON
    if args.read_scenarios:
        if not os.path.exists(input_path):
            print(f"[ERROR] 未找到需求分析文件: {input_path}", file=sys.stderr)
            sys.exit(1)
        scenarios = _read_scenarios(input_path)
        selected = [d.strip() for d in args.dimensions.split(",") if d.strip()] if args.dimensions else []
        if selected:
            scenarios = [s for s in scenarios if s["dimension_id"] in selected]
        print(json.dumps({"scenarios": scenarios}, ensure_ascii=False, indent=2))
        return

    # 模式 3/4: 写 Excel
    if not (args.write_stdin or args.write_file):
        parser.print_help(sys.stderr)
        print("\n[ERROR] 必须指定 --list / --read-scenarios / --write-stdin / --write-file 之一",
              file=sys.stderr)
        sys.exit(1)

    if args.write_stdin:
        raw = sys.stdin.read()
        try:
            data = json.loads(raw)
        except json.JSONDecodeError as e:
            print(f"[ERROR] stdin 不是合法 JSON: {e}", file=sys.stderr)
            sys.exit(1)
    else:
        if not args.json_file:
            print("[ERROR] --write-file 需要同时给 --json-file", file=sys.stderr)
            sys.exit(1)
        with open(_resolve_path(args.json_file), "r", encoding="utf-8") as f:
            try:
                data = json.load(f)
            except json.JSONDecodeError as e:
                print(f"[ERROR] json-file 不是合法 JSON: {e}", file=sys.stderr)
                sys.exit(1)

    test_cases = data.get("test_cases", []) if isinstance(data, dict) else []
    if not test_cases:
        print("[ERROR] JSON 中没有 test_cases 字段或为空", file=sys.stderr)
        sys.exit(1)

    data_output = os.path.join(data_dir, "test_cases.xlsx")
    output_path = args.output or data_output
    output_path = os.path.abspath(output_path)

    _write_excel(test_cases, output_path)
    if output_path != os.path.abspath(data_output):
        _write_excel(test_cases, os.path.abspath(data_output))

    print(f"[阶段 2/4] 测试用例 Excel 写入完成", file=sys.stderr)
    print(f"产出文件：{output_path}", file=sys.stderr)
    print(f"test cases count: {len(test_cases)}", file=sys.stderr)

    print(json.dumps({
        "stage": 2,
        "status": "ok",
        "output": output_path,
        "test_cases_count": len(test_cases),
        "by_priority": _count_by_key(test_cases, "priority"),
        "by_scenario": _count_by_key(test_cases, "scenario_id"),
    }, ensure_ascii=False, indent=2))


def _resolve_path(path):
    path = os.path.abspath(path)
    if os.path.exists(path):
        return path
    sess_out = os.getenv("SESSION_OUTPUT_DIR", "")
    if sess_out:
        alt = os.path.join(sess_out, "dataset", os.path.basename(path))
        if os.path.exists(alt):
            return alt
    return path


def _count_by_key(items, key):
    counts = {}
    for it in items:
        k = str(it.get(key, "") or "")
        counts[k] = counts.get(k, 0) + 1
    return counts


def _read_dimensions(input_path):
    from openpyxl import load_workbook
    wb = load_workbook(input_path, read_only=True)
    if "测试维度" not in wb.sheetnames:
        return []
    ws = wb["测试维度"]
    dims = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        dims.append({
            "id": str(row[0] or ""),
            "name": str(row[1] or ""),
            "type": str(row[2] or ""),
        })
    return dims


def _read_scenarios(input_path):
    from openpyxl import load_workbook
    wb = load_workbook(input_path, read_only=True)
    if "测试场景" not in wb.sheetnames:
        return []
    ws = wb["测试场景"]
    scenarios = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        scenarios.append({
            "id": str(row[0] or ""),
            "dimension_id": str(row[1] or ""),
            "name": str(row[2] or ""),
            "description": str(row[3] or ""),
        })
    return scenarios


def _write_excel(test_cases: list, path: str):
    if os.path.exists(path):
        os.remove(path)
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"
    ws.append([
        "用例 ID", "场景引用", "维度 ID", "标题", "优先级",
        "前置条件", "测试步骤", "用户输入", "预期结果", "断言类型", "状态",
    ])
    # 维度 ID 通过 scenario_id 反查（若 Agent 在 test_case 里给了 dimension_id 就用，否则留空）
    for tc in test_cases:
        steps = tc.get("steps", [])
        if isinstance(steps, list):
            steps = "\n".join(str(s) for s in steps)
        ws.append([
            tc.get("tc_id", ""),
            tc.get("scenario_id", ""),
            tc.get("dimension_id", ""),
            tc.get("title", ""),
            tc.get("priority", ""),
            tc.get("preconditions", ""),
            steps,
            tc.get("user_input", ""),
            tc.get("expected", ""),
            tc.get("assertion_type", "contains"),
            "未执行",
        ])
    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)


if __name__ == "__main__":
    try:
        main()
    except Exception:
        import traceback
        traceback.print_exc()
        sys.exit(1)
