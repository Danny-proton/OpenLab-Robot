"""
mobileAgentTest - 测试报告生成
读取需求分析、测试用例、执行结果，生成 Markdown 和 HTML 格式的测试报告。

Usage:
    python generate_report.py --requirements path/to/requirements_analysis.xlsx
                              --testcases path/to/test_cases.xlsx
                              --results path/to/execution_results.xlsx
                              --output path/to/report(.md|.html)
"""
import sys
import json
import os
import re
import argparse
import time
import html as _html_lib
from pathlib import Path


def _h(s) -> str:
    """HTML escape。"""
    return _html_lib.escape(str(s) if s is not None else "")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--requirements", default=None, help="需求分析 Excel 文件路径")
    parser.add_argument("--testcases", default=None, help="测试用例 Excel 文件路径")
    parser.add_argument("--results", default=None, help="执行结果 Excel 文件路径")
    parser.add_argument("--output", default=None, help="输出报告文件路径（.md 或 .html）")
    args = parser.parse_args()

    script_dir = os.path.dirname(os.path.abspath(__file__))
    data_dir = os.path.join(script_dir, "..", "data")

    req_path = _resolve_path(args.requirements or os.path.join(data_dir, "requirements_analysis.xlsx"))
    tc_path = _resolve_path(args.testcases or os.path.join(data_dir, "test_cases.xlsx"))
    res_path = _resolve_path(args.results or os.path.join(data_dir, "execution_results.xlsx"))
    output_base = args.output or os.path.join(data_dir, "test_report.md")

    for p, name in [(req_path, "需求分析"), (tc_path, "测试用例"), (res_path, "执行结果")]:
        if not os.path.exists(p):
            print(f"[ERROR] {name} 文件不存在: {p}", file=sys.stderr)
            print(f"      尝试路径: {p}", file=sys.stderr)
            sys.exit(1)

    dims, scenarios, dim_scenario_map = _read_requirements(req_path)
    test_cases = _read_excel_rows(tc_path)
    results = _read_excel_rows(res_path)

    if not results:
        print("[ERROR] 执行结果为空", file=sys.stderr)
        sys.exit(1)

    report_data = _build_report_data(dims, scenarios, dim_scenario_map, test_cases, results)

    md = _render_markdown(report_data)
    html = _render_html(report_data)

    ext = os.path.splitext(output_base)[1].lower()
    base_no_ext = os.path.splitext(output_base)[0]

    if ext == ".html":
        _save_output(output_base, html, "html")
        _save_output(base_no_ext + ".md", md, "md")
    else:
        _save_output(output_base, md, "md")
        _save_output(base_no_ext + ".html", html, "html")

    print(f"[阶段 4/4] 测试报告生成完成")
    print(f"产出文件：data/{os.path.basename(output_base)}")
    print(f"")
    print(f"共 {report_data['total']} 个用例，通过 {report_data['passed']}，失败 {report_data['failed']}，阻塞 {report_data['blocked']}")
    print(f"通过率：{report_data['pass_rate']:.1f}%")
    print(f"覆盖 {len(report_data['dimensions'])} 个测试维度")


def _resolve_path(path):
    path = os.path.abspath(path)
    if os.path.exists(path):
        return path
    sess_out = os.getenv("SESSION_OUTPUT_DIR", "")
    if sess_out:
        alt = os.path.join(sess_out, "dataset", os.path.basename(path))
        if os.path.exists(alt):
            return alt
    return path


def _read_excel_rows(path):
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h or "") for h in rows[0]]
    data = []
    for row in rows[1:]:
        r = {}
        for i, v in enumerate(row):
            if i < len(headers):
                r[headers[i]] = str(v or "")
        if r.get(headers[0], ""):
            data.append(r)
    return data


def _read_requirements(path):
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True)
    dims = []
    scenarios = []
    dim_scenario_map = {}

    for name in wb.sheetnames:
        if "维度" in name:
            ws = wb[name]
            rows = list(ws.iter_rows(values_only=True))
            if rows:
                h = [str(x or "") for x in rows[0]]
                for row in rows[1:]:
                    if row and row[0]:
                        dims.append({h[i]: str(row[i] or "") for i in range(len(h)) if i < len(row)})
        if "场景" in name:
            ws = wb[name]
            rows = list(ws.iter_rows(values_only=True))
            if rows:
                h = [str(x or "") for x in rows[0]]
                for row in rows[1:]:
                    if row and row[0]:
                        s = {h[i]: str(row[i] or "") for i in range(len(h)) if i < len(row)}
                        scenarios.append(s)
                        dim_id = s.get(h[1], "") if len(h) > 1 else ""
                        dim_scenario_map.setdefault(dim_id, []).append(s[list(s.keys())[0]])

    return dims, scenarios, dim_scenario_map


def _build_report_data(dims, scenarios, dim_scenario_map, test_cases, results):
    tc_map = {}
    for tc in test_cases:
        keys = list(tc.keys())
        if keys:
            tc_map[tc[keys[0]]] = tc

    scenario_map = {}
    for s in scenarios:
        keys = list(s.keys())
        if keys:
            scenario_map[s[keys[0]]] = s

    total = len(results)
    passed = sum(1 for r in results if r.get(list(r.keys())[-1], "") == "通过")
    failed = sum(1 for r in results if r.get(list(r.keys())[-1], "") == "失败")
    blocked = total - passed - failed

    total_ms = 0
    count_ms = 0
    for r in results:
        rt = r.get("响应时间", "0ms")
        digits = re.sub(r"[^\d]", "", rt)
        if digits:
            total_ms += int(digits)
            count_ms += 1
    avg_ms = total_ms / count_ms if count_ms else 0

    dim_results = {}
    for d in dims:
        d_id = d.get(list(d.keys())[0], "")
        dim_results[d_id] = {"total": 0, "passed": 0, "failed": 0, "blocked": 0, "cases": []}

    for r in results:
        keys = list(r.keys())
        tc_id = r[keys[0]]
        scenario_id = r[keys[1]] if len(keys) > 1 else ""

        sc = scenario_map.get(scenario_id, {})
        sc_keys = list(sc.keys())
        dim_id = sc[sc_keys[1]] if len(sc_keys) > 1 else ""

        tc = tc_map.get(tc_id, {})
        status = r[keys[-1]]

        rt = r.get("响应时间", "0ms")
        digits = re.sub(r"[^\d]", "", rt)

        case_info = {
            "tc_id": tc_id,
            "scenario_id": scenario_id,
            "title": r.get(keys[2], "") if len(keys) > 2 else "",
            "user_input": tc.get("用户输入", ""),
            "priority": tc.get("优先级", ""),
            "status": status,
            "response_time": rt,
            "status_code": r.get("状态码", ""),
            "expected": tc.get("预期结果", ""),
            "actual": r.get(list(r.keys())[-1], ""),
        }

        if dim_id in dim_results:
            dim_results[dim_id]["total"] += 1
            dim_results[dim_id]["cases"].append(case_info)
            if status == "通过":
                dim_results[dim_id]["passed"] += 1
            elif status == "失败":
                dim_results[dim_id]["failed"] += 1
            else:
                dim_results[dim_id]["blocked"] += 1
        else:
            dim_results.setdefault("未归类", {"total": 0, "passed": 0, "failed": 0, "blocked": 0, "cases": []})
            dim_results["未归类"]["total"] += 1
            dim_results["未归类"]["cases"].append(case_info)
            if status == "通过":
                dim_results["未归类"]["passed"] += 1
            elif status == "失败":
                dim_results["未归类"]["failed"] += 1
            else:
                dim_results["未归类"]["blocked"] += 1

    dim_list = []
    for d in dims:
        d_id = d.get(list(d.keys())[0], "")
        dr = dim_results.get(d_id, {"total": 0, "passed": 0, "failed": 0, "blocked": 0, "cases": []})
        dr["dim_id"] = d_id
        dr["dim_name"] = d.get(list(d.keys())[1] if len(list(d.keys())) > 1 else "", d_id)
        dim_list.append(dr)

    if "未归类" in dim_results:
        dim_list.append({
            "dim_id": "-",
            "dim_name": "未归类",
            "total": dim_results["未归类"]["total"],
            "passed": dim_results["未归类"]["passed"],
            "failed": dim_results["未归类"]["failed"],
            "blocked": dim_results["未归类"]["blocked"],
            "cases": dim_results["未归类"]["cases"],
        })

    pass_rate = (passed / total * 100) if total else 0

    return {
        "total": total,
        "passed": passed,
        "failed": failed,
        "blocked": blocked,
        "pass_rate": pass_rate,
        "avg_response_ms": avg_ms,
        "dimensions": dim_list,
    }


def _render_markdown(data):
    lines = []
    lines.append("# 测试执行报告")
    lines.append("")
    lines.append(f"> 生成时间：{time.strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append("")
    lines.append("---")
    lines.append("")
    lines.append("## 一、执行摘要")
    lines.append("")
    lines.append(f"- **用例总数**：{data['total']}")
    lines.append(f"- **通过**：{data['passed']} ({data['pass_rate']:.1f}%)")
    lines.append(f"- **失败**：{data['failed']} ({data['failed']/data['total']*100 if data['total'] else 0:.1f}%)")
    lines.append(f"- **阻塞**：{data['blocked']} ({data['blocked']/data['total']*100 if data['total'] else 0:.1f}%)")
    lines.append(f"- **通过率**：{data['pass_rate']:.1f}%")
    lines.append(f"- **平均响应时间**：{data['avg_response_ms']:.0f}ms")
    lines.append("")
    lines.append("```")
    bar_w = 30
    p = int(data['pass_rate'] / 100 * bar_w)
    f = int(data['failed'] / data['total'] * bar_w) if data['total'] else 0
    b = bar_w - p - f
    lines.append(f"通过:  {'█' * p}{'░' * (bar_w - p)}  {data['pass_rate']:.1f}% ({data['passed']}/{data['total']})")
    lines.append(f"失败:  {'█' * f}{'░' * (bar_w - f)}  {data['failed']/data['total']*100 if data['total'] else 0:.1f}% ({data['failed']}/{data['total']})")
    lines.append(f"阻塞:  {'█' * b}{'░' * (bar_w - b)}  {data['blocked']/data['total']*100 if data['total'] else 0:.1f}% ({data['blocked']}/{data['total']})")
    lines.append("```")
    lines.append("")
    lines.append("---")
    lines.append("")
    lines.append("## 二、按测试维度分析")
    lines.append("")

    for dim in data['dimensions']:
        dim_rate = (dim['passed'] / dim['total'] * 100) if dim['total'] else 0
        lines.append(f"### {dim['dim_name']} ({dim['dim_id']})")
        lines.append("")
        lines.append(f"- 用例数：{dim['total']} | 通过：{dim['passed']} | 失败：{dim['failed']} | 阻塞：{dim['blocked']} | 通过率：{dim_rate:.1f}%")
        lines.append("")
        lines.append("| 用例 ID | 场景 | 标题 | 用户输入 | 优先级 | 状态码 | 响应时间 | 结果 |")
        lines.append("|---------|------|------|----------|--------|--------|----------|------|")
        for c in dim['cases']:
            status_icon = {"通过": "✅", "失败": "❌", "阻塞": "⏸️"}.get(c['status'], "❓")
            lines.append(f"| {c['tc_id']} | {c['scenario_id']} | {c['title'][:30]} | {c['user_input'][:25]} | {c['priority']} | {c['status_code']} | {c['response_time']} | {status_icon} {c['status']} |")
        lines.append("")

    lines.append("---")
    lines.append("")
    lines.append("## 三、详细测试日志")
    lines.append("")
    lines.append("| 用例 ID | 场景 | 标题 | 用户输入 | 状态码 | 响应时间 | 结果 |")
    lines.append("|---------|------|------|----------|--------|----------|------|")

    all_cases = []
    for dim in data['dimensions']:
        all_cases.extend(dim['cases'])
    for c in all_cases:
        status_icon = {"通过": "✅", "失败": "❌", "阻塞": "⏸️"}.get(c['status'], "❓")
        lines.append(f"| {c['tc_id']} | {c['scenario_id']} | {c['title'][:30]} | {c['user_input'][:25]} | {c['status_code']} | {c['response_time']} | {status_icon} {c['status']} |")

    lines.append("")
    lines.append("---")
    lines.append("")
    lines.append("## 四、失败分析")
    lines.append("")
    failed_cases = [c for dim in data['dimensions'] for c in dim['cases'] if c['status'] == "失败"]
    if failed_cases:
        for c in failed_cases:
            lines.append(f"### ❌ {c['tc_id']} - {c['title']}")
            lines.append(f"- **场景**：{c['scenario_id']}")
            lines.append(f"- **优先级**：{c['priority']}")
            lines.append(f"- **状态码**：{c['status_code']}")
            lines.append(f"- **响应时间**：{c['response_time']}")
            lines.append(f"- **预期结果**：{c['expected'][:100] if c.get('expected') else 'N/A'}")
            lines.append("")
    else:
        lines.append("✅ 全部通过，无失败用例。")
        lines.append("")

    lines.append("---")
    lines.append("")
    lines.append("## 五、建议")
    lines.append("")
    if data['failed'] > 0:
        lines.append(f"- 共 {data['failed']} 个失败用例，建议优先修复高优先级失败场景。")
    if data['blocked'] > 0:
        lines.append(f"- 共 {data['blocked']} 个阻塞用例，请检查环境配置或网络连通性。")
    if data['pass_rate'] >= 90:
        lines.append("- 通过率 ≥ 90%，产品质量达标，建议进行下一轮回归测试。")
    elif data['pass_rate'] >= 70:
        lines.append("- 通过率在 70%-90% 之间，存在一定风险，建议修复后再评估。")
    else:
        lines.append("- 通过率低于 70%，产品质量存在较大风险，强烈建议修复后重新测试。")
    lines.append(f"- 平均响应时间 {data['avg_response_ms']:.0f}ms，请根据业务需求评估是否需要优化。")

    return "\n".join(lines)


def _render_html(data):
    # ------------------------------------------------------------------
    # 数据预处理：优先级分布 + 响应时间分布
    # ------------------------------------------------------------------
    priority_dist: dict[str, dict[str, int]] = {}
    rt_buckets = {"<200ms": 0, "200-500ms": 0, "500-1000ms": 0, "1-3s": 0, ">3s": 0}
    all_cases = []
    for dim in data['dimensions']:
        for c in dim['cases']:
            all_cases.append(c)
            p = c.get('priority') or '未指定'
            priority_dist.setdefault(p, {"通过": 0, "失败": 0, "阻塞": 0})
            if c['status'] in priority_dist[p]:
                priority_dist[p][c['status']] += 1
            # 响应时间分桶
            rt_str = c.get('response_time', '0ms')
            digits = re.sub(r"[^\d]", "", rt_str)
            if digits:
                ms = int(digits)
                if ms < 200: rt_buckets["<200ms"] += 1
                elif ms < 500: rt_buckets["200-500ms"] += 1
                elif ms < 1000: rt_buckets["500-1000ms"] += 1
                elif ms < 3000: rt_buckets["1-3s"] += 1
                else: rt_buckets[">3s"] += 1

    # ------------------------------------------------------------------
    # SVG 图表：维度通过率横向条形图
    # ------------------------------------------------------------------
    def _svg_dim_pass_rate(dims: list[dict]) -> str:
        if not dims:
            return '<div style="text-align:center;color:#64748b;padding:20px;font-size:12px;">无维度数据</div>'
        max_label_w = 140
        chart_w = 380
        row_h = 32
        gap = 8
        height = len(dims) * (row_h + gap) + 20
        width = max_label_w + chart_w + 80
        rows = ""
        for i, dim in enumerate(dims):
            dim_total = dim.get('total', 0) or 1
            dim_rate = dim.get('passed', 0) / dim_total * 100
            color = "#22c55e" if dim_rate >= 90 else ("#fbbf24" if dim_rate >= 70 else "#ef4444")
            y = i * (row_h + gap) + 10
            label = _h(dim.get('dim_name', dim.get('dim_id', '?'))[:14])
            bar_w_val = (dim_rate / 100) * chart_w
            rows += f"""
<g transform="translate(0,{y})">
  <text x="0" y="20" font-size="12" fill="#cbd5e1" font-family="sans-serif">{label}</text>
  <rect x="{max_label_w}" y="6" width="{chart_w}" height="{row_h - 12}" fill="rgba(15,23,42,0.6)" rx="4"/>
  <rect x="{max_label_w}" y="6" width="{bar_w_val:.1f}" height="{row_h - 12}" fill="{color}" rx="4"
    opacity="0.85" style="transition:all .3s">
    <title>{label} | 通过率: {dim_rate:.1f}% ({dim.get('passed', 0)}/{dim_total})</title>
  </rect>
  <text x="{max_label_w + bar_w_val + 8:.1f}" y="22" font-size="12" font-weight="600" fill="{color}">{dim_rate:.0f}%</text>
</g>"""
        return f'<svg width="{width}" height="{height}" viewBox="0 0 {width} {height}" style="max-width:100%">{rows}</svg>'

    # ------------------------------------------------------------------
    # SVG 图表：优先级分布饼图
    # ------------------------------------------------------------------
    def _svg_priority_pie(dist: dict[str, dict[str, int]]) -> str:
        if not dist:
            return '<div style="text-align:center;color:#64748b;padding:20px;font-size:12px;">无优先级数据</div>'
        # 简化：每个优先级一组，画堆叠柱
        items = list(dist.items())
        n = len(items)
        width = 460
        bar_w = 36
        gap = 24
        pad_l, pad_t, pad_b = 30, 30, 50
        chart_h = 160
        height = pad_t + chart_h + pad_b
        chart_w = n * (bar_w + gap)
        width = pad_l + chart_w + 30
        max_v = max((sum(d.values()) for _, d in items), default=1) or 1
        # 网格
        grid = ""
        for i in range(5):
            v = i * max_v / 4
            y = pad_t + chart_h - (v / max_v) * chart_h
            grid += f'<line x1="{pad_l}" y1="{y:.1f}" x2="{width - 20}" y2="{y:.1f}" stroke="rgba(99,102,241,0.08)"/>'
            grid += f'<text x="{pad_l - 4}" y="{y + 3:.1f}" text-anchor="end" font-size="9" fill="#64748b">{int(v)}</text>'
        bars = ""
        labels = ""
        legend = '<rect x="30" y="6" width="10" height="10" fill="#22c55e" rx="2"/><text x="44" y="14" font-size="10" fill="#cbd5e1">通过</text><rect x="90" y="6" width="10" height="10" fill="#ef4444" rx="2"/><text x="104" y="14" font-size="10" fill="#cbd5e1">失败</text><rect x="150" y="6" width="10" height="10" fill="#fbbf24" rx="2"/><text x="164" y="14" font-size="10" fill="#cbd5e1">阻塞</text>'
        for i, (prio, d) in enumerate(items):
            x = pad_l + i * (bar_w + gap)
            total = sum(d.values())
            cur_y = pad_t + chart_h
            # 通过(底) → 失败(中) → 阻塞(顶)
            for status, color in [("通过", "#22c55e"), ("失败", "#ef4444"), ("阻塞", "#fbbf24")]:
                v = d.get(status, 0)
                if v == 0: continue
                h = (v / max_v) * chart_h
                cur_y -= h
                bars += (f'<rect x="{x}" y="{cur_y:.1f}" width="{bar_w}" height="{h:.1f}" '
                         f'fill="{color}" rx="2" opacity="0.85" style="transition:all .2s">'
                         f'<title>{_h(prio)} · {status}: {v}</title></rect>')
            labels += f'<text x="{x + bar_w / 2}" y="{pad_t + chart_h + 14}" text-anchor="middle" font-size="10" fill="#94a3b8">{_h(prio[:6])}</text>'
        return f'<svg width="{width}" height="{height}" viewBox="0 0 {width} {height}" style="max-width:100%">{legend}{grid}<line x1="{pad_l}" y1="{pad_t + chart_h}" x2="{width - 20}" y2="{pad_t + chart_h}" stroke="rgba(99,102,241,0.3)"/>{bars}{labels}</svg>'

    # ------------------------------------------------------------------
    # SVG 图表：响应时间分布
    # ------------------------------------------------------------------
    def _svg_rt_distribution(buckets: dict[str, int]) -> str:
        if not any(buckets.values()):
            return '<div style="text-align:center;color:#64748b;padding:20px;font-size:12px;">无响应时间数据</div>'
        width = 460
        height = 200
        pad_l, pad_r, pad_t, pad_b = 40, 20, 30, 50
        chart_w = width - pad_l - pad_r
        chart_h = height - pad_t - pad_b
        items = list(buckets.items())
        n = len(items)
        bar_w = chart_w / n * 0.65
        gap = chart_w / n
        max_v = max(buckets.values()) or 1
        grid = ""
        for i in range(5):
            v = i * max_v / 4
            y = pad_t + chart_h - (v / max_v) * chart_h
            grid += f'<line x1="{pad_l}" y1="{y:.1f}" x2="{width - pad_r}" y2="{y:.1f}" stroke="rgba(99,102,241,0.08)"/>'
            grid += f'<text x="{pad_l - 4}" y="{y + 3:.1f}" text-anchor="end" font-size="9" fill="#64748b">{int(v)}</text>'
        bars = ""
        labels = ""
        for i, (lbl, v) in enumerate(items):
            x = pad_l + i * gap + (gap - bar_w) / 2
            h = (v / max_v) * chart_h
            y = pad_t + chart_h - h
            pct = v / sum(buckets.values()) * 100 if sum(buckets.values()) else 0
            # 颜色：随桶位变深
            colors = ["#22c55e", "#84cc16", "#fbbf24", "#f97316", "#ef4444"]
            color = colors[i % len(colors)]
            bars += (f'<rect x="{x:.1f}" y="{y:.1f}" width="{bar_w:.1f}" height="{h:.1f}" '
                     f'fill="{color}" rx="3" opacity="0.85" style="transition:all .2s">'
                     f'<title>{lbl} | 数量: {v} | 占比: {pct:.1f}%</title></rect>')
            if v > 0:
                bars += f'<text x="{x + bar_w / 2:.1f}" y="{y - 4:.1f}" text-anchor="middle" font-size="11" font-weight="600" fill="{color}">{v}</text>'
            labels += f'<text x="{x + bar_w / 2:.1f}" y="{pad_t + chart_h + 14}" text-anchor="middle" font-size="10" fill="#94a3b8">{_h(lbl)}</text>'
        return f'<svg width="{width}" height="{height}" viewBox="0 0 {width} {height}" style="max-width:100%">{grid}<line x1="{pad_l}" y1="{pad_t + chart_h}" x2="{width - pad_r}" y2="{pad_t + chart_h}" stroke="rgba(99,102,241,0.3)"/>{bars}{labels}</svg>'

    # ------------------------------------------------------------------
    # 渲染各部分
    # ------------------------------------------------------------------
    dim_rows = ""
    all_cases_html = ""
    failed_html = ""

    for dim in data['dimensions']:
        dim_rate = (dim['passed'] / dim['total'] * 100) if dim['total'] else 0
        dim_color_cls = "good" if dim_rate >= 90 else ("mid" if dim_rate >= 70 else "low")
        dim_color_hex = "#22c55e" if dim_rate >= 90 else ("#fbbf24" if dim_rate >= 70 else "#ef4444")
        bar_p = dim_rate
        bar_f = (dim['failed'] / dim['total'] * 100) if dim['total'] else 0
        bar_b = max(0, 100 - bar_p - bar_f)

        case_rows = ""
        for c in dim['cases']:
            status_cls = {"通过": "pass", "失败": "fail", "阻塞": "blocked"}.get(c['status'], "")
            case_rows += f"""<tr class="row-{status_cls}">
                <td><code>{_h(c['tc_id'])}</code></td>
                <td>{_h(c['scenario_id'])}</td>
                <td>{_h(c['title'][:40])}</td>
                <td>{_h(c['user_input'][:30])}</td>
                <td>{_h(c['priority'])}</td>
                <td>{_h(c['status_code'])}</td>
                <td class="num">{_h(c['response_time'])}</td>
                <td><span class="badge badge-{status_cls}">{_h(c['status'])}</span></td>
            </tr>\n"""
            all_cases_html += f"""<tr class="row-{status_cls}">
                <td><code>{_h(c['tc_id'])}</code></td>
                <td>{_h(c['scenario_id'])}</td>
                <td>{_h(c['title'][:40])}</td>
                <td>{_h(c['user_input'][:30])}</td>
                <td>{_h(c['status_code'])}</td>
                <td class="num">{_h(c['response_time'])}</td>
                <td><span class="badge badge-{status_cls}">{_h(c['status'])}</span></td>
            </tr>\n"""

        dim_rows += f"""
        <div class="dimension-card">
            <div class="dim-header">
                <h3><span class="dim-name">{_h(dim['dim_name'])}</span> <code class="dim-id">{_h(dim['dim_id'])}</code></h3>
                <div class="dim-stats">
                    <span class="stat stat-pass">{dim['passed']} 通过</span>
                    <span class="stat stat-fail">{dim['failed']} 失败</span>
                    <span class="stat stat-blocked">{dim['blocked']} 阻塞</span>
                    <span class="stat stat-rate {dim_color_cls}">{dim_rate:.0f}%</span>
                </div>
            </div>
            <div class="progress-bar" data-tooltip="通过 {bar_p:.1f}% / 失败 {bar_f:.1f}% / 阻塞 {bar_b:.1f}%">
                <div class="progress-pass" style="width:{bar_p:.1f}%"></div>
                <div class="progress-fail" style="width:{bar_f:.1f}%"></div>
                <div class="progress-blocked" style="width:{bar_b:.1f}%"></div>
            </div>
            <div class="table-wrap">
            <table>
                <thead><tr><th>用例 ID</th><th>场景</th><th>标题</th><th>用户输入</th><th>优先级</th><th>状态码</th><th>响应时间</th><th>结果</th></tr></thead>
                <tbody>{case_rows}</tbody>
            </table>
            </div>
        </div>"""

    for dim in data['dimensions']:
        for c in dim['cases']:
            if c['status'] == "失败":
                failed_html += f"""<div class="fail-item">
                    <h4>❌ <code>{_h(c['tc_id'])}</code> — {_h(c['title'])}</h4>
                    <p><strong>场景</strong>：{_h(c['scenario_id'])} | <strong>优先级</strong>：{_h(c['priority'])}</p>
                    <p><strong>状态码</strong>：{_h(c['status_code'])} | <strong>响应时间</strong>：{_h(c['response_time'])}</p>
                    <p><strong>预期结果</strong>：{_h((c.get('expected') or 'N/A')[:200])}</p>
                </div>"""

    bar_w = 100
    p_w = data['pass_rate']
    f_w = (data['failed'] / max(data['total'], 1)) * 100
    b_w = max(0, bar_w - p_w - f_w)

    dim_chart_svg = _svg_dim_pass_rate(data['dimensions'])
    priority_pie_svg = _svg_priority_pie(priority_dist)
    rt_dist_svg = _svg_rt_distribution(rt_buckets)

    html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>测试执行报告</title>
<style>
* {{ margin: 0; padding: 0; box-sizing: border-box; }}
@media (prefers-reduced-motion: reduce) {{
  *, *::before, *::after {{ animation: none !important; transition: none !important; }}
}}
body {{
  font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", "PingFang SC", "Microsoft YaHei", sans-serif;
  background: #0f172a;
  background-image:
    radial-gradient(at 20% 0%, rgba(99, 102, 241, 0.08) 0px, transparent 50%),
    radial-gradient(at 80% 100%, rgba(139, 92, 246, 0.06) 0px, transparent 50%);
  background-attachment: fixed;
  color: #f1f5f9; line-height: 1.6; font-size: 14px;
}}
.container {{ max-width: 1240px; margin: 0 auto; padding: 24px 16px; }}
h1 {{
  font-size: 28px; font-weight: 700;
  background: linear-gradient(135deg, #6366f1 0%, #8b5cf6 100%);
  -webkit-background-clip: text; background-clip: text;
  -webkit-text-fill-color: transparent; margin-bottom: 4px; letter-spacing: -0.5px;
}}
h2 {{
  font-size: 20px; font-weight: 700; margin: 36px 0 16px;
  border-left: 4px solid; border-image: linear-gradient(180deg, #6366f1, #8b5cf6) 1;
  padding-left: 14px; display: flex; align-items: center; gap: 8px;
}}
h3 {{ font-size: 16px; font-weight: 600; color: #f1f5f9; }}
.header {{
  background: rgba(30, 41, 59, 0.7);
  backdrop-filter: blur(12px); -webkit-backdrop-filter: blur(12px);
  border: 1px solid rgba(99, 102, 241, 0.18);
  border-radius: 16px; padding: 28px 32px; margin-bottom: 24px;
  box-shadow: 0 4px 12px rgba(0, 0, 0, 0.3); position: relative; overflow: hidden;
}}
.header::before {{
  content: ''; position: absolute; top: 0; left: 0; right: 0; height: 2px;
  background: linear-gradient(135deg, #6366f1 0%, #8b5cf6 100%); opacity: 0.8;
}}
.header h1 {{ font-size: 26px; }}
.header .time {{ font-size: 12px; color: #94a3b8; margin-top: 6px; }}
.summary-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(140px, 1fr)); gap: 12px; margin-top: 20px; }}
.summary-card {{
  background: rgba(15, 23, 42, 0.5); border: 1px solid rgba(99, 102, 241, 0.15);
  border-radius: 10px; padding: 14px; text-align: center;
  transition: all 0.2s cubic-bezier(0.4, 0, 0.2, 1); position: relative; overflow: hidden;
}}
.summary-card::before {{
  content: ''; position: absolute; left: 0; top: 0; bottom: 0; width: 3px;
  background: linear-gradient(180deg, #6366f1, #8b5cf6); opacity: 0.4; transition: all 0.2s;
}}
.summary-card:hover {{ transform: translateY(-3px); box-shadow: 0 12px 32px rgba(99, 102, 241, 0.25); border-color: rgba(99, 102, 241, 0.45); }}
.summary-card:hover::before {{ opacity: 1; width: 4px; }}
.summary-card .num {{
  font-size: 28px; font-weight: 700;
  background: linear-gradient(135deg, #6366f1 0%, #8b5cf6 100%);
  -webkit-background-clip: text; background-clip: text; -webkit-text-fill-color: transparent;
  font-variant-numeric: tabular-nums;
}}
.summary-card.pass .num {{ background: linear-gradient(135deg, #22c55e, #16a34a); -webkit-background-clip: text; background-clip: text; -webkit-text-fill-color: transparent; }}
.summary-card.fail .num {{ background: linear-gradient(135deg, #ef4444, #dc2626); -webkit-background-clip: text; background-clip: text; -webkit-text-fill-color: transparent; }}
.summary-card.blocked .num {{ background: linear-gradient(135deg, #fbbf24, #d97706); -webkit-background-clip: text; background-clip: text; -webkit-text-fill-color: transparent; }}
.summary-card.rate .num {{ background: linear-gradient(135deg, #8b5cf6, #a78bfa); -webkit-background-clip: text; background-clip: text; -webkit-text-fill-color: transparent; }}
.summary-card .label {{ font-size: 11px; color: #64748b; margin-top: 4px; text-transform: uppercase; letter-spacing: 0.5px; }}
.progress-bar {{
  height: 14px; background: rgba(15, 23, 42, 0.6); border-radius: 8px;
  overflow: hidden; display: flex; margin: 10px 0 14px;
  border: 1px solid rgba(99, 102, 241, 0.15); position: relative;
}}
.progress-bar:hover {{ box-shadow: 0 0 0 2px rgba(99, 102, 241, 0.25); }}
.progress-pass {{ background: linear-gradient(135deg, #22c55e, #16a34a); transition: width 0.6s; }}
.progress-fail {{ background: linear-gradient(135deg, #ef4444, #dc2626); transition: width 0.6s; }}
.progress-blocked {{ background: linear-gradient(135deg, #fbbf24, #d97706); transition: width 0.6s; }}
.chart-section {{
  background: rgba(30, 41, 59, 0.6);
  backdrop-filter: blur(12px); -webkit-backdrop-filter: blur(12px);
  border: 1px solid rgba(99, 102, 241, 0.15); border-radius: 12px;
  padding: 20px 24px; margin: 16px 0;
}}
.chart-title {{ font-size: 13px; color: #cbd5e1; font-weight: 600; margin-bottom: 12px; }}
.chart-hint {{ font-size: 11px; color: #64748b; margin-top: 8px; text-align: center; }}
.charts-row {{ display: grid; grid-template-columns: 1fr 1fr; gap: 16px; }}
@media (max-width: 820px) {{ .charts-row {{ grid-template-columns: 1fr; }} }}
.dimension-card {{
  background: rgba(30, 41, 59, 0.6);
  backdrop-filter: blur(12px); -webkit-backdrop-filter: blur(12px);
  border: 1px solid rgba(99, 102, 241, 0.15); border-radius: 12px;
  padding: 20px 24px; margin-bottom: 16px;
  transition: all 0.2s; position: relative;
}}
.dimension-card:hover {{ border-color: rgba(99, 102, 241, 0.35); transform: translateY(-2px); box-shadow: 0 8px 24px rgba(0,0,0,0.3); }}
.dim-header {{ display: flex; justify-content: space-between; align-items: center; flex-wrap: wrap; gap: 8px; margin-bottom: 4px; }}
.dim-id {{ font-size: 11px; color: #94a3b8; font-weight: 400; padding: 2px 6px; background: rgba(15,23,42,0.6); border-radius: 4px; }}
.dim-stats {{ display: flex; gap: 10px; font-size: 12px; }}
.stat {{ padding: 2px 10px; border-radius: 999px; font-weight: 600; }}
.stat-pass {{ color: #22c55e; background: rgba(34,197,94,0.12); border: 1px solid rgba(34,197,94,0.3); }}
.stat-fail {{ color: #ef4444; background: rgba(239,68,68,0.12); border: 1px solid rgba(239,68,68,0.3); }}
.stat-blocked {{ color: #fbbf24; background: rgba(251,191,36,0.12); border: 1px solid rgba(251,191,36,0.3); }}
.stat-rate {{ font-weight: 700; font-size: 14px; padding: 2px 12px; }}
.stat-rate.good {{ color: #22c55e; background: rgba(34,197,94,0.12); }}
.stat-rate.mid {{ color: #fbbf24; background: rgba(251,191,36,0.12); }}
.stat-rate.low {{ color: #ef4444; background: rgba(239,68,68,0.12); }}
.table-wrap {{ overflow-x: auto; border-radius: 8px; margin-top: 8px; }}
table {{ width: 100%; border-collapse: collapse; font-size: 12px; }}
th {{
  background: rgba(15, 23, 42, 0.6); color: #94a3b8; text-align: left;
  padding: 10px 12px; font-weight: 600; font-size: 11px;
  text-transform: uppercase; letter-spacing: 0.5px;
  border-bottom: 1px solid rgba(99, 102, 241, 0.15);
}}
td {{ padding: 10px 12px; border-bottom: 1px solid rgba(99, 102, 241, 0.08); color: #cbd5e1; }}
td.num {{ text-align: right; font-variant-numeric: tabular-nums; }}
tbody tr {{ transition: all 0.15s; }}
tbody tr:hover {{ background: rgba(99, 102, 241, 0.08); box-shadow: inset 2px 0 0 #6366f1; }}
tbody tr.row-fail td {{ background: rgba(239, 68, 68, 0.06); }}
tbody tr.row-fail:hover td {{ background: rgba(239, 68, 68, 0.12); box-shadow: inset 2px 0 0 #ef4444; }}
tbody tr.row-blocked td {{ background: rgba(251, 191, 36, 0.06); }}
tbody tr.row-blocked:hover td {{ background: rgba(251, 191, 36, 0.12); box-shadow: inset 2px 0 0 #fbbf24; }}
code {{ background: rgba(15,23,42,0.6); padding: 2px 6px; border-radius: 4px; font-size: 0.88em; color: #c7d2fe; font-family: "SF Mono", Monaco, monospace; }}
.badge {{ display: inline-block; padding: 2px 10px; border-radius: 999px; font-size: 11px; font-weight: 600; letter-spacing: 0.3px; }}
.badge-pass {{ background: rgba(34,197,94,0.15); color: #22c55e; border: 1px solid rgba(34,197,94,0.3); }}
.badge-fail {{ background: rgba(239,68,68,0.15); color: #ef4444; border: 1px solid rgba(239,68,68,0.3); }}
.badge-blocked {{ background: rgba(251,191,36,0.15); color: #fbbf24; border: 1px solid rgba(251,191,36,0.3); }}
.fail-item {{
  background: rgba(239, 68, 68, 0.06); border-left: 4px solid #ef4444;
  padding: 14px 18px; margin-bottom: 12px; border-radius: 0 8px 8px 0;
  transition: all 0.2s;
}}
.fail-item:hover {{ transform: translateX(2px); background: rgba(239, 68, 68, 0.12); }}
.fail-item h4 {{ color: #ef4444; font-size: 14px; margin-bottom: 6px; font-weight: 600; }}
.fail-item p {{ font-size: 12px; color: #cbd5e1; margin: 2px 0; }}
.fail-item strong {{ color: #f1f5f9; font-weight: 600; }}
.section {{
  background: rgba(30, 41, 59, 0.6);
  backdrop-filter: blur(12px); -webkit-backdrop-filter: blur(12px);
  border: 1px solid rgba(99, 102, 241, 0.15); border-radius: 12px;
  padding: 20px 24px;
}}
.footer {{ margin-top: 40px; padding-top: 16px; border-top: 1px solid rgba(99, 102, 241, 0.15); color: #64748b; font-size: 12px; text-align: center; }}
</style>
</head>
<body>
<div class="container">
    <div class="header">
        <h1>📊 测试执行报告</h1>
        <div class="time">生成时间：{time.strftime('%Y-%m-%d %H:%M:%S')}</div>
        <div class="summary-grid">
            <div class="summary-card total"><div class="num">{data['total']}</div><div class="label">用例总数</div></div>
            <div class="summary-card pass"><div class="num">{data['passed']}</div><div class="label">通过</div></div>
            <div class="summary-card fail"><div class="num">{data['failed']}</div><div class="label">失败</div></div>
            <div class="summary-card blocked"><div class="num">{data['blocked']}</div><div class="label">阻塞</div></div>
            <div class="summary-card rate"><div class="num">{data['pass_rate']:.1f}%</div><div class="label">通过率</div></div>
        </div>
        <div class="progress-bar" style="margin-top:18px;height:16px" data-tooltip="通过 {p_w:.1f}% / 失败 {f_w:.1f}% / 阻塞 {b_w:.1f}%">
            <div class="progress-pass" style="width:{p_w:.1f}%"></div>
            <div class="progress-fail" style="width:{f_w:.1f}%"></div>
            <div class="progress-blocked" style="width:{b_w:.1f}%"></div>
        </div>
        <div style="margin-top:8px;font-size:12px;color:#94a3b8">平均响应时间：{data['avg_response_ms']:.0f}ms</div>
    </div>

    <h2>📈 可视化分析</h2>
    <div class="charts-row">
        <div class="chart-section">
            <div class="chart-title">各维度通过率（hover 条形查看明细）</div>
            <div style="display:flex;justify-content:center;overflow-x:auto;">{dim_chart_svg}</div>
            <div class="chart-hint">绿色 ≥90% · 黄色 70-90% · 红色 &lt;70%</div>
        </div>
        <div class="chart-section">
            <div class="chart-title">优先级 × 状态堆叠分布（hover 各段查看数量）</div>
            <div style="display:flex;justify-content:center;overflow-x:auto;">{priority_pie_svg}</div>
            <div class="chart-hint">每柱代表一个优先级，堆叠通过/失败/阻塞计数</div>
        </div>
    </div>
    <div class="chart-section">
        <div class="chart-title">响应时间分布（hover 柱查看数量与占比）</div>
        <div style="display:flex;justify-content:center;overflow-x:auto;">{rt_dist_svg}</div>
        <div class="chart-hint">按响应时间分桶统计，颜色由绿到红代表时延递增</div>
    </div>

    <h2>📂 按测试维度分析</h2>
    {dim_rows}

    <h2>📋 详细测试日志</h2>
    <div class="section">
    <div class="table-wrap">
    <table>
        <thead><tr><th>用例 ID</th><th>场景</th><th>标题</th><th>用户输入</th><th>状态码</th><th>响应时间</th><th>结果</th></tr></thead>
        <tbody>{all_cases_html}</tbody>
    </table>
    </div>
    </div>

    <h2>🔍 失败分析</h2>
    <div class="section">
    {failed_html if failed_html else '<p style="padding:20px;text-align:center;color:#22c55e;font-size:16px">✅ 全部通过，无失败用例。</p>'}
    </div>

    <h2>💡 建议</h2>
    <div class="section">
    <ul style="padding-left:20px;color:#cbd5e1;font-size:13px;line-height:2;">
        {'<li>共 <strong style="color:#ef4444">' + str(data['failed']) + '</strong> 个失败用例，建议优先修复高优先级失败场景。</li>' if data['failed'] > 0 else ''}
        {'<li>共 <strong style="color:#fbbf24">' + str(data['blocked']) + '</strong> 个阻塞用例，请检查环境配置或网络连通性。</li>' if data['blocked'] > 0 else ''}
        <li>{'✅ <span style="color:#22c55e">通过率 ≥ 90%</span>，产品质量达标，建议进行下一轮回归测试。' if data['pass_rate'] >= 90 else '⚠️ <span style="color:#fbbf24">通过率在 70%-90% 之间</span>，存在一定风险，建议修复后再评估。' if data['pass_rate'] >= 70 else '❌ <span style="color:#ef4444">通过率低于 70%</span>，产品质量存在较大风险，强烈建议修复后重新测试。'}</li>
        <li>平均响应时间 <strong>{data['avg_response_ms']:.0f}ms</strong>，请根据业务需求评估是否需要优化。</li>
    </ul>
    </div>

    <div class="footer">本报告由 agent-eval-v1.1.1 generate_report.py 自动生成 · 深色玻璃态可视化</div>
</div>
</body>
</html>"""
    return html


def _save_output(path, content, mode):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        f.write(content)
    print(f"[OK] {mode.upper()} report saved: {path}", file=sys.stderr)


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        import traceback
        traceback.print_exc()
        sys.exit(1)
