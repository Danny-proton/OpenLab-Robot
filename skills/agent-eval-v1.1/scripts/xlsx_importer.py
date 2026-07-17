#!/usr/bin/env python3
"""xlsx_importer.py — Excel 导入/导出标注模板 + 用例导入。

用法:
  python xlsx_importer.py --config .agent-eval/config.yaml --export-annotations
  python xlsx_importer.py --config .agent-eval/config.yaml --import-annotations input.xlsx
  python xlsx_importer.py --config .agent-eval/config.yaml --import-cases input.xlsx --split train
"""

from __future__ import annotations

import argparse
import csv
import json
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

sys.path.insert(0, str(Path(__file__).resolve().parent))
import common as C  # noqa: E402

try:
    from openpyxl import Workbook, load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# 3+N 子指标列名
THREE_N_COLUMNS = [
    "bv_decision_accuracy",
    "bv_sop_completeness",
    "bv_customer_experience",
    "bv_terminology_accuracy",
    "rc_risk_classification",
]


def export_annotations_xlsx(cfg: C.EvalConfig, out_path: Path) -> None:
    """导出标注模板 XLSX。"""
    if not HAS_OPENPYXL:
        _export_annotations_csv(cfg, out_path)
        return

    from annotator import load_annotations, load_annotations_by_case, THREE_N_SUB_METRICS
    cases = C.load_yaml(cfg.cases_dir / "train.yaml").get("cases", [])
    by_case = load_annotations_by_case(cfg)

    wb = Workbook()

    # Sheet1: 用例基本信息
    ws1 = wb.active
    ws1.title = "用例信息"
    ws1.append(["case_id", "name", "prompt", "risk_level", "dimension"])
    for case in cases:
        ws1.append([
            case.get("id", ""),
            case.get("name", ""),
            case.get("prompt", "")[:200],
            case.get("risk_level", ""),
            ", ".join(case.get("dimensions", []) or []),
        ])

    # Sheet2: Ground Truth
    ws2 = wb.create_sheet("Ground Truth")
    ws2.append(["case_id", "decision", "risk_level", "sop_complete"])
    for case in cases:
        cid = case.get("id", "")
        ann = (by_case.get(cid, []) or [{}])[-1] if cid in by_case else {}
        gt = ann.get("ground_truth", {})
        ws2.append([
            cid,
            gt.get("decision", ""),
            gt.get("risk_level", ""),
            "是" if gt.get("sop_complete") else ("否" if gt.get("sop_complete") is False else ""),
        ])

    # Sheet3: 3+N 指标打分
    ws3 = wb.create_sheet("3+N指标打分")
    ws3.append(["case_id"] + THREE_N_SUB_METRICS)
    for case in cases:
        cid = case.get("id", "")
        ann = (by_case.get(cid, []) or [{}])[-1] if cid in by_case else {}
        ql = ann.get("quality_labels", {})
        row = [cid]
        for m in THREE_N_SUB_METRICS:
            row.append(ql.get(m, ""))
        ws3.append(row)

    # Sheet4: 人工判定
    ws4 = wb.create_sheet("人工判定")
    ws4.append(["case_id", "pass/fail/疑问", "理由"])
    for case in cases:
        cid = case.get("id", "")
        ann = (by_case.get(cid, []) or [{}])[-1] if cid in by_case else {}
        ql = ann.get("quality_labels", {})
        hf = ann.get("human_feedback", {})
        verdict = ""
        if ql.get("pass") is True:
            verdict = "pass"
        elif ql.get("pass") is False:
            verdict = "fail"
        else:
            verdict = "疑问"
        ws4.append([
            cid,
            verdict,
            ql.get("reason", "") or hf.get("suggestion", ""),
        ])

    wb.save(str(out_path))
    print(f"[xlsx_importer] 标注模板已导出: {out_path}")


def _export_annotations_csv(cfg: C.EvalConfig, out_path: Path) -> None:
    """CSV 降级导出（无 openpyxl 时）。"""
    from annotator import load_annotations_by_case
    cases = C.load_yaml(cfg.cases_dir / "train.yaml").get("cases", [])
    by_case = load_annotations_by_case(cfg)

    csv_path = out_path.with_suffix(".csv")
    with csv_path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["case_id", "name", "risk_level", "dimension",
                          "decision", "sop_complete",
                          "bv_decision_accuracy", "bv_sop_completeness",
                          "bv_customer_experience", "bv_terminology_accuracy",
                          "rc_risk_classification", "pass_fail", "reason"])
        for case in cases:
            cid = case.get("id", "")
            ann = (by_case.get(cid, []) or [{}])[-1] if cid in by_case else {}
            gt = ann.get("ground_truth", {})
            ql = ann.get("quality_labels", {})
            writer.writerow([
                cid, case.get("name", ""),
                case.get("risk_level", ""), ", ".join(case.get("dimensions", []) or []),
                gt.get("decision", ""), gt.get("sop_complete", ""),
                ql.get("bv_decision_accuracy", ""), ql.get("bv_sop_completeness", ""),
                ql.get("bv_customer_experience", ""), ql.get("bv_terminology_accuracy", ""),
                ql.get("rc_risk_classification", ""),
                "pass" if ql.get("pass") else ("fail" if ql.get("pass") is False else "疑问"),
                ql.get("reason", ""),
            ])
    print(f"[xlsx_importer] CSV 降级导出: {csv_path}")


def import_annotations_xlsx(cfg: C.EvalConfig, xlsx_path: Path) -> tuple[int, list[str]]:
    """导入标注结果，返回 (成功数, 错误列表)。"""
    if not HAS_OPENPYXL:
        return _import_annotations_csv(cfg, xlsx_path.with_suffix(".csv"))

    from annotator import save_annotation, validate_annotation
    wb = load_workbook(str(xlsx_path))

    # 读取各 Sheet
    # Sheet2: Ground Truth
    gt_data = {}
    ws2 = wb.get_sheet_by_name("Ground Truth")
    if ws2:
        for row in ws2.iter_rows(min_row=2, values_only=True):
            if row[0]:
                gt_data[row[0]] = {
                    "decision": str(row[1] or ""),
                    "risk_level": str(row[2] or ""),
                    "sop_complete": row[3] == "是" if row[3] else None,
                }

    # Sheet3: 3+N 指标打分
    ql_data = {}
    ws3 = wb.get_sheet_by_name("3+N指标打分")
    if ws3:
        for row in ws3.iter_rows(min_row=2, values_only=True):
            if row[0]:
                ql = {}
                for i, m in enumerate(THREE_N_COLUMNS):
                    val = row[i + 1] if len(row) > i + 1 else None
                    if val is not None:
                        try:
                            ql[m] = int(val)
                        except (ValueError, TypeError):
                            pass
                ql_data[row[0]] = ql

    # Sheet4: 人工判定
    hf_data = {}
    ws4 = wb.get_sheet_by_name("人工判定")
    if ws4:
        for row in ws4.iter_rows(min_row=2, values_only=True):
            if row[0]:
                verdict = str(row[1] or "").strip().lower()
                if verdict == "pass":
                    passed = True
                elif verdict == "fail":
                    passed = False
                else:
                    passed = None
                hf_data[row[0]] = {
                    "pass": passed,
                    "reason": str(row[2] or ""),
                }

    # 合并并保存
    n_ok, n_err = 0, 0
    errors = []
    all_case_ids = set(gt_data.keys()) | set(ql_data.keys()) | set(hf_data.keys())

    for cid in all_case_ids:
        ann = {
            "run_id": f"import_{C.now_iso().replace(':', '-').replace('+', '-')[:19]}",
            "case_id": cid,
            "annotator": "excel_import",
            "timestamp": C.now_iso(),
            "ground_truth": gt_data.get(cid, {}),
            "quality_labels": {**ql_data.get(cid, {}), **hf_data.get(cid, {})},
            "human_feedback": {"approved": hf_data.get(cid, {}).get("pass"), "suggestion": hf_data.get(cid, {}).get("reason", "")},
        }

        errs = validate_annotation(ann)
        if errs:
            n_err += 1
            errors.append(f"case={cid}: {errs}")
        else:
            save_annotation(cfg, ann)
            n_ok += 1

    return n_ok, errors


def _import_annotations_csv(cfg: C.EvalConfig, csv_path: Path) -> tuple[int, list[str]]:
    """CSV 降级导入。"""
    from annotator import save_annotation, validate_annotation
    n_ok, n_err = 0, 0
    errors = []

    with csv_path.open("r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            cid = row.get("case_id", "")
            if not cid:
                continue
            ql = {}
            for m in THREE_N_COLUMNS:
                val = row.get(m, "")
                if val:
                    try:
                        ql[m] = int(val)
                    except (ValueError, TypeError):
                        pass

            verdict = row.get("pass_fail", "").strip().lower()
            passed = True if verdict == "pass" else (False if verdict == "fail" else None)

            ann = {
                "run_id": f"import_{C.now_iso().replace(':', '-').replace('+', '-')[:19]}",
                "case_id": cid,
                "annotator": "csv_import",
                "timestamp": C.now_iso(),
                "ground_truth": {
                    "decision": row.get("decision", ""),
                    "risk_level": row.get("risk_level", ""),
                },
                "quality_labels": {
                    **ql,
                    "pass": passed,
                    "reason": row.get("reason", ""),
                },
                "human_feedback": {"approved": passed, "suggestion": row.get("reason", "")},
            }
            errs = validate_annotation(ann)
            if errs:
                n_err += 1
                errors.append(f"case={cid}: {errs}")
            else:
                save_annotation(cfg, ann)
                n_ok += 1

    return n_ok, errors


def import_cases_xlsx(cfg: C.EvalConfig, xlsx_path: Path, split: str) -> tuple[int, list[str]]:
    """从 Excel 导入用例，转换为 YAML。"""
    if not HAS_OPENPYXL:
        print("[xlsx_importer] 需要 openpyxl 库来导入 Excel: pip install openpyxl")
        return 0, ["missing openpyxl"]

    from annotator import infer_risk_and_dimension
    wb = load_workbook(str(xlsx_path))
    ws = wb.active

    cases = []
    errors = []
    seen_ids = set()

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0]:
            continue

        case_id = str(row[0]).strip()
        if not case_id:
            errors.append(f"行 {i}: case_id 为空")
            continue
        if case_id in seen_ids:
            errors.append(f"行 {i}: case_id 重复: {case_id}")
            continue
        seen_ids.add(case_id)

        name = str(row[1] or "").strip()
        prompt = str(row[2] or "").strip()
        if not prompt:
            errors.append(f"行 {i}: prompt 为空")
            continue

        case: dict[str, Any] = {
            "id": case_id,
            "name": name,
            "prompt": prompt,
        }

        # 风险等级和维度
        risk_level = str(row[3] or "").strip().lower() if len(row) > 3 else ""
        dimensions = [d.strip() for d in str(row[4] or "").split(",") if d.strip()] if len(row) > 4 else []

        if not risk_level or not dimensions:
            inferred_risk, inferred_dims = infer_risk_and_dimension(case)
            case["risk_level"] = risk_level or inferred_risk
            case["dimensions"] = dimensions or inferred_dims
        else:
            case["risk_level"] = risk_level
            case["dimensions"] = dimensions

        cases.append(case)

    # 输出 YAML
    out_path = cfg.cases_dir / f"imported_{datetime.now().strftime('%Y%m%d')}.yaml"
    C.dump_yaml({"cases": cases}, out_path)
    print(f"[xlsx_importer] 导入 {len(cases)} 条用例到 {out_path}")
    if errors:
        for e in errors:
            print(f"  [错误] {e}")
    return len(cases), errors


def main() -> int:
    ap = argparse.ArgumentParser(description="Excel 导入/导出")
    ap.add_argument("--config", required=True)
    ap.add_argument("--export-annotations", action="store_true", help="导出标注模板 XLSX")
    ap.add_argument("--import-annotations", metavar="FILE", help="导入标注结果")
    ap.add_argument("--import-cases", metavar="FILE", help="从 Excel 导入用例")
    ap.add_argument("--split", default="train", help="用例 split")
    ap.add_argument("--out", default=None, help="输出路径")
    args = ap.parse_args()

    cfg = C.EvalConfig.load(Path(args.config).resolve())

    if args.export_annotations:
        out = Path(args.out) if args.out else cfg.root / "annotations_template.xlsx"
        export_annotations_xlsx(cfg, out)
        return 0

    if args.import_annotations:
        xlsx_path = Path(args.import_annotations)
        if not xlsx_path.exists():
            sys.stderr.write(f"文件不存在: {xlsx_path}\n")
            return 2
        n_ok, errs = import_annotations_xlsx(cfg, xlsx_path)
        print(f"[xlsx_importer] 导入完成: 成功 {n_ok}, 失败 {len(errs)}")
        for e in errs:
            print(f"  [错误] {e}")
        return 0 if not errs else 1

    if args.import_cases:
        xlsx_path = Path(args.import_cases)
        if not xlsx_path.exists():
            sys.stderr.write(f"文件不存在: {xlsx_path}\n")
            return 2
        n, errs = import_cases_xlsx(cfg, xlsx_path, args.split)
        return 0

    ap.print_help()
    return 0


if __name__ == "__main__":
    sys.exit(main())